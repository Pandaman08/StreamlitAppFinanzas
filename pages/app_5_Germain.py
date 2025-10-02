import streamlit as st
import pandas as pd
import io
from bs4 import BeautifulSoup
import re
import unicodedata
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import ColorChoice

st.set_page_config(page_title="Consolidador SMV - Finanzas Corporativas", layout="wide")

# Hola
# ================= HEADER =================
st.title("📊 Consolidador de Estados Financieros - SMV")
st.markdown("**Análisis Financiero Automatizado** | Sube archivos Excel del SMV y obtén análisis completo con gráficas.")

# ================= SIDEBAR =================
with st.sidebar:
    st.header("⚙️ Configuración")
    nombre_empresa = st.text_input("Nombre de la Empresa", value="EMPRESA ANALIZADA", help="Aparecerá en el reporte")
    
    st.markdown("---")
    st.markdown("### 📋 Instrucciones")
    st.info("""
    1. Descarga los 5 archivos Excel (2020-2024) del SMV
    2. Súbelos todos a la vez
    3. Espera el procesamiento
    4. Revisa resultados y descarga el consolidado
    """)

# ================= UPLOAD FILES =================
archivos = st.file_uploader(
    "📁 Selecciona los 5 archivos Excel (.xls) - años 2020 a 2024",
    type=["xls"],
    accept_multiple_files=True
)

if not archivos:
    st.warning("👆 **Por favor, sube los archivos Excel del SMV para comenzar el análisis.**")
    st.stop()

# ================= UTILIDADES =================
def normalize_name(s):
    """Normaliza textos: quita tildes, mayúsculas, compacta espacios y elimina notas (9)"""
    if not isinstance(s, str):
        return s
    s2 = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('ASCII')
    s2 = re.sub(r'\s+', ' ', s2).strip().upper()
    s2 = re.sub(r'\s*\(\d+\)\s*$', '', s2)
    return s2

def limpiar_valor(valor):
    """Limpia y convierte strings numéricos a float. Maneja paréntesis como negativos."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    if not valor or valor == '0' or valor == '':
        return 0.0
    valor = str(valor).replace(',', '').replace('\xa0', '').replace(' ', '')
    if valor.startswith('(') and valor.endswith(')'):
        valor = '-' + valor[1:-1]
    try:
        return float(valor)
    except:
        return 0.0

def buscar_cuenta_exacta(df, keywords):
    """Busca una cuenta que contenga todas las palabras clave"""
    for idx in df.index:
        if all(kw.lower() in idx.lower() for kw in keywords):
            return idx
    return None

# ================= PROCESAR ARCHIVOS =================
datos_balance = {}
datos_resultados = {}
datos_flujo_efectivo = {}
archivos_procesados = []

progress_bar = st.progress(0)
status_text = st.empty()

for i, archivo in enumerate(archivos):
    status_text.text(f"📦 Procesando: {archivo.name}")
    
    contenido = None
    for cod in ['latin-1', 'cp1252', 'utf-8']:
        try:
            archivo.seek(0)
            contenido = archivo.read().decode(cod)
            break
        except:
            continue
    
    if not contenido:
        st.error(f"❌ No se pudo leer {archivo.name}")
        continue
    
    soup = BeautifulSoup(contenido, 'html.parser')
    
    # Procesar Balance
    tabla_balance = soup.find('table', {'id': 'gvReporte'})
    if tabla_balance:
        filas = [[td.get_text(strip=True) for td in tr.find_all(['td', 'th'])] for tr in tabla_balance.find_all('tr') if tr.find_all(['td', 'th'])]
        
        if len(filas) > 1:
            encabezados = filas[0]
            columnas_anios = encabezados[2:]
            anios = [int(m.group(0)) if (m := re.search(r'\b(19|20)\d{2}\b', col)) else None for col in columnas_anios]
            
            current_section = None
            for fila in filas[1:]:
                if len(fila) < 2:
                    continue
                first_cell = fila[0].strip()
                if re.search(r'ACTIVO|PASIVO|PATRIMONIO|TOTAL', first_cell, flags=re.IGNORECASE) and len(fila) <= 2:
                    current_section = normalize_name(first_cell)
                    continue
                
                cuenta_raw = fila[0]
                for i_col, valor_str in enumerate(fila[2:]):
                    anio = anios[i_col]
                    if anio is None or anio < 2020:
                        continue
                    valor = limpiar_valor(valor_str)
                    cuenta = normalize_name(cuenta_raw)
                    cuenta_key = f"{current_section}||{cuenta}" if current_section else cuenta
                    
                    if anio not in datos_balance:
                        datos_balance[anio] = {}
                    
                    if cuenta_key in datos_balance[anio]:
                        existing = datos_balance[anio][cuenta_key]
                        if existing == 0 and valor != 0:
                            datos_balance[anio][cuenta_key] = valor
                    else:
                        datos_balance[anio][cuenta_key] = valor
    
    # Procesar Estado de Resultados
    tabla_resultados = soup.find('table', {'id': 'gvReporte1'})
    if tabla_resultados:
        filas = [[td.get_text(strip=True) for td in tr.find_all(['td', 'th'])] for tr in tabla_resultados.find_all('tr') if tr.find_all(['td', 'th'])]
        
        if len(filas) > 1:
            encabezados = filas[0]
            columnas_anios = encabezados[2:]
            anios = [int(m.group(0)) if (m := re.search(r'\b(19|20)\d{2}\b', col)) else None for col in columnas_anios]
            
            for fila in filas[1:]:
                if len(fila) < 2:
                    continue
                cuenta_raw = fila[0]
                for i_col, valor_str in enumerate(fila[2:]):
                    anio = anios[i_col]
                    if anio is None or anio < 2020:
                        continue
                    valor = limpiar_valor(valor_str)
                    cuenta = normalize_name(cuenta_raw)
                    
                    if anio not in datos_resultados:
                        datos_resultados[anio] = {}
                    
                    if cuenta in datos_resultados[anio]:
                        existing = datos_resultados[anio][cuenta]
                        if existing == 0 and valor != 0:
                            datos_resultados[anio][cuenta] = valor
                    else:
                        datos_resultados[anio][cuenta] = valor
    
    # Procesar Estado de Flujo de Efectivo
    tabla_flujo = soup.find('table', {'id': 'gvReporte2'})
    if tabla_flujo:
        filas = [[td.get_text(strip=True) for td in tr.find_all(['td', 'th'])] for tr in tabla_flujo.find_all('tr') if tr.find_all(['td', 'th'])]
        
        if len(filas) > 1:
            encabezados = filas[0]
            columnas_anios = encabezados[2:]
            anios = [int(m.group(0)) if (m := re.search(r'\b(19|20)\d{2}\b', col)) else None for col in columnas_anios]
            
            for fila in filas[1:]:
                if len(fila) < 2:
                    continue
                cuenta_raw = fila[0]
                for i_col, valor_str in enumerate(fila[2:]):
                    anio = anios[i_col]
                    if anio is None or anio < 2020:
                        continue
                    valor = limpiar_valor(valor_str)
                    cuenta = normalize_name(cuenta_raw)
                    
                    if anio not in datos_flujo_efectivo:
                        datos_flujo_efectivo[anio] = {}
                    
                    if cuenta in datos_flujo_efectivo[anio]:
                        existing = datos_flujo_efectivo[anio][cuenta]
                        if existing == 0 and valor != 0:
                            datos_flujo_efectivo[anio][cuenta] = valor
                    else:
                        datos_flujo_efectivo[anio][cuenta] = valor
    
    archivos_procesados.append(archivo.name)
    progress_bar.progress((i + 1) / len(archivos))

status_text.empty()
progress_bar.empty()

# ================= CREAR DATAFRAMES =================
df_balance = pd.DataFrame.from_dict(datos_balance, orient='index').fillna(0.0).T if datos_balance else pd.DataFrame()
df_resultados = pd.DataFrame.from_dict(datos_resultados, orient='index').fillna(0.0).T if datos_resultados else pd.DataFrame()
df_flujo_efectivo = pd.DataFrame.from_dict(datos_flujo_efectivo, orient='index').fillna(0.0).T if datos_flujo_efectivo else pd.DataFrame()

if not df_balance.empty:
    df_balance.index = [normalize_name(i) for i in df_balance.index]
    df_balance = df_balance.reindex(sorted(df_balance.columns), axis=1)

if not df_resultados.empty:
    df_resultados.index = [normalize_name(i) for i in df_resultados.index]
    df_resultados = df_resultados.reindex(sorted(df_resultados.columns), axis=1)

if not df_flujo_efectivo.empty:
    df_flujo_efectivo.index = [normalize_name(i) for i in df_flujo_efectivo.index]
    df_flujo_efectivo = df_flujo_efectivo.reindex(sorted(df_flujo_efectivo.columns), axis=1)

# ================= ANÁLISIS VERTICAL Y HORIZONTAL =================
# BALANCE
df_vertical_balance = pd.DataFrame()
df_horizontal_balance = pd.DataFrame()

if not df_balance.empty:
    df_vertical_balance = df_balance.copy()
    total_activos_row = None
    for idx in df_vertical_balance.index:
        if "TOTAL" in idx and ("ACTIVOS" in idx or "ACTIVO" in idx) and "CORRIENTE" not in idx:
            total_activos_row = idx
            break
    
    if total_activos_row:
        total_activos = df_vertical_balance.loc[total_activos_row]
        for col in df_vertical_balance.columns:
            if total_activos[col] != 0:
                df_vertical_balance[col] = (df_vertical_balance[col] / total_activos[col]) * 100
        df_vertical_balance = df_vertical_balance.round(1)
    
    # Horizontal
    df_horizontal_balance = df_balance.copy()
    columnas = df_horizontal_balance.columns.tolist()
    nuevas_columnas = []
    for i in range(len(columnas) - 1):
        anio_actual = columnas[i + 1]
        anio_anterior = columnas[i]
        col_nombre = f"{anio_anterior}-{anio_actual}"
        df_horizontal_balance[col_nombre] = (df_horizontal_balance[anio_actual] - df_horizontal_balance[anio_anterior]) / df_horizontal_balance[anio_anterior] * 100
        nuevas_columnas.append(col_nombre)
    df_horizontal_balance = df_horizontal_balance[nuevas_columnas].round(1)
    df_horizontal_balance = df_horizontal_balance.replace([float('inf'), float('-inf')], pd.NA)

# ESTADO DE RESULTADOS
df_vertical_resultados = pd.DataFrame()
df_horizontal_resultados = pd.DataFrame()

if not df_resultados.empty:
    df_vertical_resultados = df_resultados.copy()
    ventas_row = buscar_cuenta_exacta(df_resultados, ["INGRESOS", "ACTIVIDADES", "ORDINARIAS"]) or buscar_cuenta_exacta(df_resultados, ["VENTAS", "NETAS"])
    
    if ventas_row:
        ventas = df_vertical_resultados.loc[ventas_row]
        for col in df_vertical_resultados.columns:
            if ventas[col] != 0:
                df_vertical_resultados[col] = (df_vertical_resultados[col] / ventas[col]) * 100
        df_vertical_resultados = df_vertical_resultados.round(1)
    
    # Horizontal
    df_horizontal_resultados = df_resultados.copy()
    columnas = df_horizontal_resultados.columns.tolist()
    nuevas_columnas = []
    for i in range(len(columnas) - 1):
        anio_actual = columnas[i + 1]
        anio_anterior = columnas[i]
        col_nombre = f"{anio_anterior}-{anio_actual}"
        df_horizontal_resultados[col_nombre] = (df_horizontal_resultados[anio_actual] - df_horizontal_resultados[anio_anterior]) / df_horizontal_resultados[anio_anterior] * 100
        nuevas_columnas.append(col_nombre)
    df_horizontal_resultados = df_horizontal_resultados[nuevas_columnas].round(1)
    df_horizontal_resultados = df_horizontal_resultados.replace([float('inf'), float('-inf')], pd.NA)

# ================= CÁLCULO DE RATIOS =================
ratios_data = {}
anios_comunes = sorted(list(set(df_balance.columns) & set(df_resultados.columns))) if (not df_balance.empty and not df_resultados.empty) else []

if anios_comunes:
    for i, anio in enumerate(anios_comunes):
        ratios_data[anio] = {}
        
        # Balance
        act_corr = buscar_cuenta_exacta(df_balance, ["TOTAL", "ACTIVO", "CORRIENTE"])
        activo_corriente = df_balance.loc[act_corr, anio] if act_corr else 0.0
        
        inv1 = buscar_cuenta_exacta(df_balance, ["INVENTARIOS"])
        inv2 = buscar_cuenta_exacta(df_balance, ["EXISTENCIAS"])
        inventarios = df_balance.loc[inv1 or inv2, anio] if (inv1 or inv2) else 0.0
        
        pas_corr = buscar_cuenta_exacta(df_balance, ["TOTAL", "PASIVO", "CORRIENTE"])
        pasivo_corriente = df_balance.loc[pas_corr, anio] if pas_corr else 0.0
        
        cxc1 = buscar_cuenta_exacta(df_balance, ["CUENTAS", "COBRAR", "COMERCIALES"])
        cxc2 = buscar_cuenta_exacta(df_balance, ["CUENTAS", "COBRAR", "VINCULADAS"])
        cxc3 = buscar_cuenta_exacta(df_balance, ["OTRAS", "CUENTAS", "COBRAR"])
        cxc_val = sum(df_balance.loc[cxc, anio] for cxc in [cxc1, cxc2, cxc3] if cxc and cxc in df_balance.index)
        
        act_tot = buscar_cuenta_exacta(df_balance, ["TOTAL", "ACTIVO"])
        activos_totales = df_balance.loc[act_tot, anio] if act_tot else 0.0
        
        pas_tot = buscar_cuenta_exacta(df_balance, ["TOTAL", "PASIVO"])
        pasivo_total = df_balance.loc[pas_tot, anio] if pas_tot else 0.0
        
        # Patrimonio - búsqueda mejorada
        patrimonio = 0.0
        patr_variantes = [
            ["TOTAL", "PATRIMONIO"],
            ["PATRIMONIO", "NETO"],
            ["TOTAL", "PATRIMONIO", "NETO"],
        ]
        
        for variante in patr_variantes:
            patr_found = buscar_cuenta_exacta(df_balance, variante)
            if patr_found:
                patrimonio = df_balance.loc[patr_found, anio]
                break
        
        if patrimonio == 0.0:
            for idx in df_balance.index:
                if "PATRIMONIO" in idx and "TOTAL" in idx:
                    patrimonio = df_balance.loc[idx, anio]
                    break
        
        if patrimonio == 0.0 and activos_totales != 0.0:
            patrimonio = activos_totales - pasivo_total
        
        # Estado de Resultados
        ventas1 = buscar_cuenta_exacta(df_resultados, ["INGRESOS", "ACTIVIDADES", "ORDINARIAS"])
        ventas2 = buscar_cuenta_exacta(df_resultados, ["VENTAS", "NETAS"])
        ventas = df_resultados.loc[ventas1 or ventas2, anio] if (ventas1 or ventas2) else 0.0
        
        costo1 = buscar_cuenta_exacta(df_resultados, ["COSTO", "VENTAS"])
        costo_ventas = df_resultados.loc[costo1, anio] if costo1 else 0.0
        
        # Utilidad Neta - búsqueda mejorada
        utilidad_neta = 0.0
        util_variantes = [
            ["UTILIDAD", "NETA", "EJERCICIO"],
            ["GANANCIA", "NETA", "EJERCICIO"],
            ["UTILIDAD", "PERDIDA", "NETA"],
            ["GANANCIA", "PERDIDA", "NETA"],
        ]
        
        for variante in util_variantes:
            util_found = buscar_cuenta_exacta(df_resultados, variante)
            if util_found:
                utilidad_neta = df_resultados.loc[util_found, anio]
                break
        
        if utilidad_neta == 0.0:
            for idx in df_resultados.index:
                if "NETA" in idx and "EJERCICIO" in idx:
                    utilidad_neta = df_resultados.loc[idx, anio]
                    break
        
        # Promedios
        cxc_prom = cxc_val
        inv_prom = inventarios
        act_prom = activos_totales
        patr_prom = patrimonio
        
        if i > 0:
            anio_ant = anios_comunes[i-1]
            cxc_ant = sum(df_balance.loc[cxc, anio_ant] for cxc in [cxc1, cxc2, cxc3] if cxc and cxc in df_balance.index)
            inv_ant = df_balance.loc[inv1 or inv2, anio_ant] if (inv1 or inv2) else 0.0
            act_ant = df_balance.loc[act_tot, anio_ant] if act_tot else 0.0
            
            patr_ant = 0.0
            for variante in patr_variantes:
                patr_found = buscar_cuenta_exacta(df_balance, variante)
                if patr_found:
                    patr_ant = df_balance.loc[patr_found, anio_ant]
                    break
            
            if patr_ant == 0.0:
                for idx in df_balance.index:
                    if "PATRIMONIO" in idx and "TOTAL" in idx:
                        patr_ant = df_balance.loc[idx, anio_ant]
                        break
            
            if patr_ant == 0.0 and act_ant != 0.0:
                pas_ant = df_balance.loc[pas_tot, anio_ant] if pas_tot else 0.0
                patr_ant = act_ant - pas_ant
            
            cxc_prom = (cxc_val + cxc_ant) / 2
            inv_prom = (inventarios + inv_ant) / 2
            act_prom = (activos_totales + act_ant) / 2
            patr_prom = (patrimonio + patr_ant) / 2
        
        # Calcular ratios
        ratios_data[anio]["Liquidez Corriente"] = activo_corriente / pasivo_corriente if pasivo_corriente != 0 else None
        ratios_data[anio]["Prueba Ácida"] = (activo_corriente - inventarios) / pasivo_corriente if pasivo_corriente != 0 else None
        ratios_data[anio]["Rotación CxC"] = ventas / cxc_prom if cxc_prom != 0 else None
        ratios_data[anio]["Rotación Inventarios"] = abs(costo_ventas) / inv_prom if inv_prom != 0 else None
        ratios_data[anio]["Rotación Activos Totales"] = ventas / act_prom if act_prom != 0 else None
        ratios_data[anio]["Razón Deuda Total"] = pasivo_total / activos_totales if activos_totales != 0 else None
        ratios_data[anio]["Razón Deuda/Patrimonio"] = pasivo_total / patrimonio if patrimonio != 0 else None
        ratios_data[anio]["Margen Neto"] = utilidad_neta / ventas if ventas != 0 else None
        ratios_data[anio]["ROA"] = utilidad_neta / act_prom if act_prom != 0 else None
        ratios_data[anio]["ROE"] = utilidad_neta / patr_prom if patr_prom != 0 else None

df_ratios = pd.DataFrame.from_dict(ratios_data, orient='index').round(4).T if ratios_data else pd.DataFrame()

# ================= SIDEBAR STATUS =================
with st.sidebar:
    st.markdown("---")
    st.success(f"✅ **{len(archivos_procesados)}** archivos procesados")
    if anios_comunes:
        st.info(f"📅 **Años:** {', '.join(map(str, anios_comunes))}")
    st.metric("Ratios Calculados", len(df_ratios) if not df_ratios.empty else 0)

# ================= TABS =================
tab1, tab2, tab3, tab4 = st.tabs(["📊 Estados Financieros", "📈 Análisis V/H", "🧮 Ratios y Gráficas", "📥 Descargar"])

with tab1:
    st.subheader("💼 Estado de Situación Financiera")
    if not df_balance.empty:
        st.dataframe(df_balance, use_container_width=True)
    else:
        st.warning("No se encontró data del Balance")
    
    st.markdown("---")
    st.subheader("💰 Estado de Resultados")
    if not df_resultados.empty:
        st.dataframe(df_resultados, use_container_width=True)
    else:
        st.warning("No se encontró data del Estado de Resultados")
    
    st.markdown("---")
    st.subheader("💵 Estado de Flujo de Efectivo")
    if not df_flujo_efectivo.empty:
        st.dataframe(df_flujo_efectivo, use_container_width=True)
    else:
        st.warning("No se encontró data del Flujo de Efectivo")

with tab2:
    st.subheader("📊 Análisis Vertical y Horizontal - Estado de Situación Financiera")
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Análisis Vertical (%)**")
        if not df_vertical_balance.empty:
            st.dataframe(df_vertical_balance, use_container_width=True)
    
    with col2:
        st.markdown("**Análisis Horizontal (Variación %)**")
        if not df_horizontal_balance.empty:
            st.dataframe(df_horizontal_balance, use_container_width=True)
    
    st.markdown("---")
    st.subheader("📊 Análisis Vertical y Horizontal - Estado de Resultados")
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Análisis Vertical (%)**")
        if not df_vertical_resultados.empty:
            st.dataframe(df_vertical_resultados, use_container_width=True)
    
    with col2:
        st.markdown("**Análisis Horizontal (Variación %)**")
        if not df_horizontal_resultados.empty:
            st.dataframe(df_horizontal_resultados, use_container_width=True)

with tab3:
    st.subheader("🧮 Ratios Financieros")
    
    if not df_ratios.empty:
        # Métricas principales
        ultimo_anio = df_ratios.columns[-1]
        penultimo_anio = df_ratios.columns[-2] if len(df_ratios.columns) > 1 else ultimo_anio
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            valor_actual = df_ratios.loc['ROE', ultimo_anio] if 'ROE' in df_ratios.index and not pd.isna(df_ratios.loc['ROE', ultimo_anio]) else 0
            valor_anterior = df_ratios.loc['ROE', penultimo_anio] if 'ROE' in df_ratios.index and not pd.isna(df_ratios.loc['ROE', penultimo_anio]) else 0
            delta = valor_actual - valor_anterior
            st.metric("ROE", f"{valor_actual:.2%}", delta=f"{delta:.2%}")
        
        with col2:
            valor_actual = df_ratios.loc['ROA', ultimo_anio] if 'ROA' in df_ratios.index and not pd.isna(df_ratios.loc['ROA', ultimo_anio]) else 0
            valor_anterior = df_ratios.loc['ROA', penultimo_anio] if 'ROA' in df_ratios.index and not pd.isna(df_ratios.loc['ROA', penultimo_anio]) else 0
            delta = valor_actual - valor_anterior
            st.metric("ROA", f"{valor_actual:.2%}", delta=f"{delta:.2%}")
        
        with col3:
            valor_actual = df_ratios.loc['Liquidez Corriente', ultimo_anio] if 'Liquidez Corriente' in df_ratios.index and not pd.isna(df_ratios.loc['Liquidez Corriente', ultimo_anio]) else 0
            valor_anterior = df_ratios.loc['Liquidez Corriente', penultimo_anio] if 'Liquidez Corriente' in df_ratios.index and not pd.isna(df_ratios.loc['Liquidez Corriente', penultimo_anio]) else 0
            delta = valor_actual - valor_anterior
            st.metric("Liquidez Corriente", f"{valor_actual:.2f}", delta=f"{delta:.2f}")
        
        with col4:
            valor_actual = df_ratios.loc['Margen Neto', ultimo_anio] if 'Margen Neto' in df_ratios.index and not pd.isna(df_ratios.loc['Margen Neto', ultimo_anio]) else 0
            valor_anterior = df_ratios.loc['Margen Neto', penultimo_anio] if 'Margen Neto' in df_ratios.index and not pd.isna(df_ratios.loc['Margen Neto', penultimo_anio]) else 0
            delta = valor_actual - valor_anterior
            st.metric("Margen Neto", f"{valor_actual:.2%}", delta=f"{delta:.2%}")
        
        st.markdown("---")
        
        # Tabla de ratios
        st.markdown("### 📋 Tabla de Ratios")
        st.dataframe(df_ratios, use_container_width=True)
        
        st.markdown("---")
        
        # Mapa de Calor
        st.markdown("### 🔥 Mapa de Calor - Todos los Ratios")
        fig_heatmap = go.Figure(data=go.Heatmap(
            z=df_ratios.values,
            x=df_ratios.columns,
            y=df_ratios.index,
            colorscale='RdYlGn',
            text=df_ratios.round(2).values,
            texttemplate='%{text}',
            textfont={"size": 10},
            colorbar=dict(title="Valor")
        ))
        fig_heatmap.update_layout(
            title=f"Mapa de Calor de Ratios Financieros - {nombre_empresa}",
            xaxis_title="Año",
            yaxis_title="Ratio",
            height=500
        )
        st.plotly_chart(fig_heatmap, use_container_width=True)
        
        st.markdown("---")
        st.markdown("### 📈 Gráficas Individuales por Ratio")
        
        # Crear gráficas individuales en 2 columnas
        col1, col2 = st.columns(2)
        
        for idx, ratio in enumerate(df_ratios.index):
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=df_ratios.columns,
                y=df_ratios.loc[ratio],
                mode='lines+markers',
                name=ratio,
                line=dict(width=3),
                marker=dict(size=10)
            ))
            fig.update_layout(
                title=f"{ratio}",
                xaxis_title="Año",
                yaxis_title="Valor",
                height=350,
                showlegend=False
            )
            
            if idx % 2 == 0:
                with col1:
                    st.plotly_chart(fig, use_container_width=True)
            else:
                with col2:
                    st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning("No se pudieron calcular ratios")

with tab4:
    st.subheader("📥 Descargar Reporte Consolidado")
    st.markdown(f"**Empresa:** {nombre_empresa}")
    st.markdown(f"**Años analizados:** {', '.join(map(str, anios_comunes)) if anios_comunes else 'N/A'}")
    st.info("📊 El archivo Excel incluirá todas las gráficas de ratios financieros")
    
    # Crear Excel con formato profesional
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not df_balance.empty:
            df_balance.to_excel(writer, sheet_name='Balance', index_label='Cuenta')
        if not df_resultados.empty:
            df_resultados.to_excel(writer, sheet_name='Estado Resultados', index_label='Cuenta')
        if not df_flujo_efectivo.empty:
            df_flujo_efectivo.to_excel(writer, sheet_name='Flujo Efectivo', index_label='Cuenta')
        
        # Análisis Balance (V y H en la misma hoja)
        if not df_vertical_balance.empty and not df_horizontal_balance.empty:
            startrow = 0
            df_vertical_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta', startrow=startrow)
            startrow = len(df_vertical_balance) + 3
            ws = writer.sheets['Analisis Balance']
            ws.cell(row=startrow, column=1, value="ANÁLISIS HORIZONTAL (Variación %)")
            df_horizontal_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta', startrow=startrow+1, header=True)
        elif not df_vertical_balance.empty:
            df_vertical_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta')
        elif not df_horizontal_balance.empty:
            df_horizontal_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta')
        
        # Análisis Estado de Resultados (V y H en la misma hoja)
        if not df_vertical_resultados.empty and not df_horizontal_resultados.empty:
            startrow = 0
            df_vertical_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta', startrow=startrow)
            startrow = len(df_vertical_resultados) + 3
            ws = writer.sheets['Analisis Resultados']
            ws.cell(row=startrow, column=1, value="ANÁLISIS HORIZONTAL (Variación %)")
            df_horizontal_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta', startrow=startrow+1, header=True)
        elif not df_vertical_resultados.empty:
            df_vertical_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta')
        elif not df_horizontal_resultados.empty:
            df_horizontal_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta')
        
        if not df_ratios.empty:
            df_ratios.to_excel(writer, sheet_name='Ratios', index_label='Ratio')
    
    # Aplicar formato profesional y agregar gráficas
    output.seek(0)
    wb = load_workbook(output)
    
    # Estilo profesional
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(name='Calibri', size=10, bold=True)
    cell_font = Font(name='Calibri', size=10)
    subtitle_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    subtitle_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Formato de encabezados
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Buscar subtítulos (ANÁLISIS HORIZONTAL)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if isinstance(cell.value, str) and "ANÁLISIS HORIZONTAL" in cell.value:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=cell.row, column=col).fill = subtitle_fill
                        ws.cell(row=cell.row, column=col).font = subtitle_font
                        ws.cell(row=cell.row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=cell.row + 1, column=col).fill = header_fill
                        ws.cell(row=cell.row + 1, column=col).font = header_font
                        ws.cell(row=cell.row + 1, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        # Formato de celdas y totales
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                
                if isinstance(cell.value, str) and "TOTAL" in cell.value.upper():
                    for c in row:
                        c.fill = total_fill
                        c.font = total_font
                
                if isinstance(cell.value, (int, float)) and cell.column > 1:
                    if 'Analisis' in sheet_name:
                        cell.number_format = '0.0"%"'
                    elif 'Ratios' in sheet_name:
                        cell.number_format = '0.0000'
                    else:
                        cell.number_format = '#,##0'
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # ================= AGREGAR GRÁFICAS AL EXCEL =================
    if not df_ratios.empty and 'Ratios' in wb.sheetnames:
        ws_ratios = wb['Ratios']
        
        # Crear hoja nueva para gráficas y mapa de calor juntos
        if 'Ratios y Graficas' in wb.sheetnames:
            del wb['Ratios y Graficas']
        ws_graficas = wb.create_sheet('Ratios y Graficas')
        
        # ===== SECCIÓN 1: MAPA DE CALOR =====
        ws_graficas['A1'] = 'MAPA DE CALOR DE RATIOS FINANCIEROS'
        ws_graficas['A1'].font = Font(name='Calibri', size=14, bold=True, color="366092")
        
        ws_graficas.append([])
        ws_graficas.append(['Ratio / Año'] + df_ratios.columns.tolist())
        for ratio in df_ratios.index:
            row_data = [ratio] + df_ratios.loc[ratio].tolist()
            ws_graficas.append(row_data)
        
        from openpyxl.formatting.rule import ColorScaleRule
        
        num_years = len(df_ratios.columns)
        ratio_names = df_ratios.index.tolist()
        
        color_scale = ColorScaleRule(
            start_type='min', start_color='F8696B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'
        )
        
        ws_graficas.conditional_formatting.add(
            f'B3:{get_column_letter(num_years+1)}{len(ratio_names)+2}',
            color_scale
        )
        
        for cell in ws_graficas[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for row in ws_graficas.iter_rows(min_row=3, max_row=len(ratio_names)+2, min_col=1, max_col=num_years+1):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.column > 1:
                    cell.number_format = '0.0000'
        
        ws_graficas.column_dimensions['A'].width = 30
        for col in range(2, num_years + 2):
            ws_graficas.column_dimensions[get_column_letter(col)].width = 12
        
        # ===== SECCIÓN 2: GRÁFICAS INDIVIDUALES =====
        chart_start_row = len(ratio_names) + 5
        ws_graficas.cell(row=chart_start_row, column=1, value='GRÁFICAS INDIVIDUALES POR RATIO')
        ws_graficas.cell(row=chart_start_row, column=1).font = Font(name='Calibri', size=14, bold=True, color="366092")
        
        chart_row = chart_start_row + 2
        chart_col = 1
        charts_per_row = 2
        chart_height = 15
        chart_width = 10
        
        for idx, ratio in enumerate(ratio_names):
            chart = LineChart()
            chart.title = ratio
            chart.style = 10
            chart.y_axis.title = "Valor"
            chart.x_axis.title = "Año"
            chart.height = 7
            chart.width = 12
            
            ratio_row = ratio_names.index(ratio) + 2
            
            data = Reference(ws_ratios, min_col=2, min_row=ratio_row, max_col=num_years+1, max_row=ratio_row)
            cats = Reference(ws_ratios, min_col=2, min_row=1, max_col=num_years+1, max_row=1)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            
            row_position = chart_row + (idx // charts_per_row) * chart_height
            col_position = chart_col + (idx % charts_per_row) * chart_width
            
            ws_graficas.add_chart(chart, f"{get_column_letter(col_position)}{row_position}")
    
    # Guardar Excel formateado
    output_formatted = io.BytesIO()
    wb.save(output_formatted)
    output_formatted.seek(0)
    
    st.download_button(
        label="📥 Descargar Excel Consolidado (Con Gráficas)",
        data=output_formatted.getvalue(),
        file_name=f"Analisis_Financiero_{nombre_empresa.replace(' ', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.success("✅ ¡Listo! El archivo incluye todos los análisis con gráficas integradas en Excel.")