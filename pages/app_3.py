import streamlit as st
import pandas as pd
import io
from bs4 import BeautifulSoup
import re
import unicodedata
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference,Series,ScatterChart
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart.axis import ChartLines
import matplotlib
matplotlib.use('Agg')  # Para evitar problemas en Streamlit
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
import os
import tempfile

st.set_page_config(page_title="Consolidador SMV - Finanzas Corporativas", layout="wide")

# ================= HEADER =================
st.title("📊 Consolidador de Estados Financieros - SMV")
st.markdown("**Análisis Financiero Automatizado** | Sube archivos Excel del SMV (2002-2024) y obtén análisis completo con gráficas.")

# ================= SIDEBAR =================
with st.sidebar:
    st.header("⚙️ Configuración")
    nombre_empresa = st.text_input("Nombre de la Empresa", value="EMPRESA ANALIZADA", help="Aparecerá en el reporte")
    
    st.markdown("---")
    st.markdown("### 📋 Instrucciones")
    st.info("""
    1. Descarga archivos Excel (.xls) del SMV
    2. Súbelos (pueden ser de cualquier año: 2002-2024)
    3. Espera el procesamiento
    4. Revisa resultados y descarga el consolidado
    """)

# ================= UPLOAD FILES =================
archivos = st.file_uploader(
    "📁 Selecciona archivos Excel (.xls) del SMV",
    type=["xls"],
    accept_multiple_files=True
)

if not archivos:
    st.warning("👆 **Por favor, sube los archivos Excel del SMV para comenzar el análisis.**")
    st.stop()

# ================= UTILIDADES =================
def normalize_name(s):
    """Normaliza textos: quita tildes, mayúsculas, compacta espacios y elimina notas (9)"""
    if not isinstance(s, str):
        return s
    s2 = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('ASCII')
    s2 = re.sub(r'\s+', ' ', s2).strip().upper()
    s2 = re.sub(r'\s*\(\d+\)\s*$', '', s2)
    return s2

def limpiar_valor(valor):
    """Limpia y convierte strings numéricos a float. Maneja paréntesis como negativos."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    if not valor or valor == '0' or valor == '':
        return 0.0
    valor = str(valor).replace(',', '').replace('\xa0', '').replace(' ', '')
    if valor.startswith('(') and valor.endswith(')'):
        valor = '-' + valor[1:-1]
    try:
        return float(valor)
    except:
        return 0.0

def mapear_cuenta_normalizada(cuenta_original, anio):
    """Mapea nombres de cuentas antiguas (pre-2010) a nomenclatura moderna."""
    cuenta = normalize_name(cuenta_original)
    
    mapeo_antiguo = {
        # Balance - Activos
        "CAJA Y BANCOS": "EFECTIVO Y EQUIVALENTES AL EFECTIVO",
        "VALORES NEGOCIABLES": "OTROS ACTIVOS FINANCIEROS",
        "EXISTENCIAS": "INVENTARIOS",
        "GASTOS PAGADOS POR ANTICIPADO": "ANTICIPOS",
        "INVERSIONES PERMANENTES": "INVERSIONES EN SUBSIDIARIAS NEGOCIOS CONJUNTOS Y ASOCIADAS",
        "INMUEBLES MAQUINARIA Y EQUIPO NETO DE DEPRECIACION ACUMULADA": "PROPIEDADES PLANTA Y EQUIPO",
        "INMUEBLES MAQUINARIA Y EQUIPO": "PROPIEDADES PLANTA Y EQUIPO",
        "ACTIVO INTANGIBLE NETO DE DEPRECIACION ACUMULADA": "ACTIVOS INTANGIBLES DISTINTOS DE LA PLUSVALIA",
        "ACTIVOS INTANGIBLES": "ACTIVOS INTANGIBLES DISTINTOS DE LA PLUSVALIA",
        "OTROS ACTIVOS": "OTROS ACTIVOS NO FINANCIEROS",
        "IMPUESTO A LA RENTA Y PARTICIPACIONES DIFERIDOS ACTIVO": "ACTIVOS POR IMPUESTOS DIFERIDOS",
        
        # Balance - Pasivos
        "SOBREGIROS Y PAGARES BANCARIOS": "OTROS PASIVOS FINANCIEROS",
        "PARTE CORRIENTE DE LAS DEUDAS A LARGO PLAZO": "OTROS PASIVOS FINANCIEROS",
        "DEUDAS A LARGO PLAZO": "OTROS PASIVOS FINANCIEROS",
        "INGRESOS DIFERIDOS": "INGRESOS DIFERIDOS",
        "IMPUESTO A LA RENTA Y PARTICIPACIONES DIFERIDOS PASIVO": "PASIVOS POR IMPUESTOS DIFERIDOS",
        
        # Balance - Patrimonio
        "CAPITAL": "CAPITAL EMITIDO",
        "CAPITAL ADICIONAL": "PRIMAS DE EMISION",
        "EXCEDENTE DE REVALUACION": "SUPERAVIT DE REVALUACION",
        "RESERVAS LEGALES": "OTRAS RESERVAS DE CAPITAL",
        "OTRAS RESERVAS": "OTRAS RESERVAS DE PATRIMONIO",
        "RESULTADOS ACUMULADOS": "RESULTADOS ACUMULADOS",
        
        # Estado de Resultados
        "VENTAS NETAS INGRESOS OPERACIONALES": "INGRESOS DE ACTIVIDADES ORDINARIAS",
        "VENTAS NETAS": "INGRESOS DE ACTIVIDADES ORDINARIAS",
        "OTROS INGRESOS OPERACIONALES": "OTROS INGRESOS OPERATIVOS",
        "TOTAL DE INGRESOS BRUTOS": "INGRESOS DE ACTIVIDADES ORDINARIAS",
        "COSTO DE VENTAS": "COSTO DE VENTAS",
        "UTILIDAD BRUTA": "GANANCIA PERDIDA BRUTA",
        "GASTOS DE ADMINISTRACION": "GASTOS DE ADMINISTRACION",
        "GASTOS DE VENTAS": "GASTOS DE VENTAS Y DISTRIBUCION",
        "UTILIDAD OPERATIVA": "GANANCIA PERDIDA OPERATIVA",
        "INGRESOS FINANCIEROS": "INGRESOS FINANCIEROS",
        "GASTOS FINANCIEROS": "GASTOS FINANCIEROS",
        "OTROS INGRESOS": "OTROS INGRESOS OPERATIVOS",
        "OTROS GASTOS": "OTROS GASTOS OPERATIVOS",
        "RESULTADO POR EXPOSICION A LA INFLACION": "DIFERENCIAS DE CAMBIO NETO",
        "RESULTADOS ANTES DE PARTIDAS EXTRAORDINARIAS PARTICIPACIONES Y DEL IMPUESTO A LA RENTA": "GANANCIA PERDIDA ANTES DE IMPUESTOS",
        "PARTICIPACIONES": "OTROS INGRESOS GASTOS DE LAS SUBSIDIARIAS ASOCIADAS Y NEGOCIOS CONJUNTOS",
        "IMPUESTO A LA RENTA": "INGRESO GASTO POR IMPUESTO",
        "RESULTADO ANTES DE PARTIDAS EXTRAORDINARIAS": "GANANCIA PERDIDA NETA DE OPERACIONES CONTINUADAS",
        "INGRESOS EXTRAORDINARIOS": "OTROS INGRESOS OPERATIVOS",
        "GASTOS EXTRAORDINARIOS": "OTROS GASTOS OPERATIVOS",
        "RESULTADO ANTES DE INTERES MINORITARIO": "GANANCIA PERDIDA NETA DEL EJERCICIO",
        "INTERES MINORITARIO": "PARTICIPACION NO CONTROLADORA",
        "UTILIDAD PERDIDA NETA DEL EJERCICIO": "GANANCIA PERDIDA NETA DEL EJERCICIO",
        "UTILIDAD NETA DEL EJERCICIO": "GANANCIA PERDIDA NETA DEL EJERCICIO",
        "UTILIDAD PERDIDA NETA ATRIBUIBLE A LOS ACCIONISTAS": "GANANCIA PERDIDA NETA DEL EJERCICIO"
    }
    
    if anio < 2010:
        if cuenta in mapeo_antiguo:
            return mapeo_antiguo[cuenta]
        
        # Búsqueda flexible
        if "VENTAS" in cuenta and "NETAS" in cuenta:
            return "INGRESOS DE ACTIVIDADES ORDINARIAS"
        if "UTILIDAD" in cuenta and "NETA" in cuenta and "EJERCICIO" in cuenta:
            return "GANANCIA PERDIDA NETA DEL EJERCICIO"
        if "EXISTENCIAS" in cuenta:
            return "INVENTARIOS"
    
    return cuenta

def buscar_cuenta_flexible(df, keywords_list):
    """Busca una cuenta que coincida con cualquiera de las listas de keywords."""
    for keywords in keywords_list:
        for idx in df.index:
            if all(kw.upper() in idx.upper() for kw in keywords):
                return idx
    return None

def buscar_cuenta_parcial(df, keywords):
    """Búsqueda con coincidencia parcial (al menos una palabra clave)."""
    for idx in df.index:
        if any(kw.upper() in idx.upper() for kw in keywords):
            return idx
    return None

# ================= PROCESAR ARCHIVOS =================
datos_balance = {}
datos_resultados = {}
datos_flujo_efectivo = {}

progress_bar = st.progress(0)
status_text = st.empty()

for i, archivo in enumerate(archivos):
    status_text.text(f"📦 Procesando: {archivo.name}")
    
    contenido = None
    for cod in ['latin-1', 'cp1252', 'utf-8']:
        try:
            archivo.seek(0)
            contenido = archivo.read().decode(cod)
            break
        except:
            continue
    
    if not contenido:
        st.error(f"❌ No se pudo leer {archivo.name}")
        continue
    
    soup = BeautifulSoup(contenido, 'html.parser')
    
    # Procesar Balance
    tabla_balance = soup.find('table', {'id': 'gvReporte'})
    
    if tabla_balance:
        filas = []
        for tr in tabla_balance.find_all('tr'):
            celdas = [td.get_text(strip=True) for td in tr.find_all(['td', 'th'])]
            if celdas:
                filas.append(celdas)

        if len(filas) > 1:
            encabezados = filas[0]
            columnas_anios = encabezados[2:]
            anios = []
            for col in columnas_anios:
                m = re.search(r'\b(19|20)\d{2}\b', col)
                if m:
                    anios.append(int(m.group(0)))
                else:
                    anios.append(None)

            encabezados_seccion = [
                "ACTIVOS", "ACTIVO", "ACTIVOS CORRIENTES", "ACTIVO CORRIENTE",
                "ACTIVOS NO CORRIENTES", "ACTIVO NO CORRIENTE",
                "PASIVOS", "PASIVO", "PASIVOS CORRIENTES", "PASIVO CORRIENTE",
                "PASIVOS NO CORRIENTES", "PASIVO NO CORRIENTE",
                "PATRIMONIO", "PATRIMONIO NETO", "PASIVO Y PATRIMONIO", "PASIVOS Y PATRIMONIO",
                "CUENTAS POR COBRAR COMERCIALES Y OTRAS CUENTAS POR COBRAR",
                "CUENTAS POR PAGAR COMERCIALES Y OTRAS CUENTAS POR PAGAR"
            ]

            for fila in filas[1:]:
                if len(fila) < 3:
                    continue
                
                cuenta_raw = fila[0].strip()
                
                if not cuenta_raw:
                    continue
                
                cuenta_normalizada_temp = normalize_name(cuenta_raw)
                
                if cuenta_normalizada_temp in encabezados_seccion:
                    continue
                
                valores_fila = [limpiar_valor(v) for v in fila[2:]]
                if all(v == 0 for v in valores_fila):
                    continue
                
                for i_col, valor_str in enumerate(fila[2:]):
                    anio = anios[i_col]
                    if anio is None:
                        continue
                    
                    valor = limpiar_valor(valor_str)
                    cuenta_normalizada = mapear_cuenta_normalizada(cuenta_raw, anio)
                    
                    if anio not in datos_balance:
                        datos_balance[anio] = {}
                    
                    if cuenta_normalizada not in datos_balance[anio]:
                        datos_balance[anio][cuenta_normalizada] = valor
                    elif datos_balance[anio][cuenta_normalizada] == 0 and valor != 0:
                        datos_balance[anio][cuenta_normalizada] = valor

    # Procesar Estado de Resultados
    tabla_resultados = soup.find('table', {'id': 'gvReporte1'})
    
    if tabla_resultados:
        filas = []
        for tr in tabla_resultados.find_all('tr'):
            celdas = [td.get_text(strip=True) for td in tr.find_all(['td', 'th'])]
            if celdas:
                filas.append(celdas)

        if len(filas) > 1:
            encabezados = filas[0]
            columnas_anios = encabezados[2:]
            anios = []
            
            for col in columnas_anios:
                m = re.search(r'\b(19|20)\d{2}\b', col)
                if m:
                    anios.append(int(m.group(0)))
                else:
                    anios.append(None)

            for fila in filas[1:]:
                if len(fila) < 2:
                    continue
                
                cuenta_raw = fila[0].strip()
                
                if len(fila) <= 2:
                    continue
                
                for i_col, valor_str in enumerate(fila[2:]):
                    anio = anios[i_col]
                    if anio is None:
                        continue
                    
                    valor = limpiar_valor(valor_str)
                    cuenta_normalizada = mapear_cuenta_normalizada(cuenta_raw, anio)
                    
                    if anio not in datos_resultados:
                        datos_resultados[anio] = {}
                    
                    datos_resultados[anio][cuenta_normalizada] = valor

    # Procesar Flujo de Efectivo
    tabla_flujo = soup.find('table', {'id': 'gvReporte3'})
    
    if tabla_flujo:
        filas = []
        for tr in tabla_flujo.find_all('tr'):
            celdas = [td.get_text(strip=True) for td in tr.find_all(['td', 'th'])]
            if celdas:
                filas.append(celdas)

        if len(filas) > 1:
            encabezados = filas[0]
            columnas_anios = encabezados[2:]
            anios = []
            
            for col in columnas_anios:
                m = re.search(r'\b(19|20)\d{2}\b', col)
                if m:
                    anios.append(int(m.group(0)))
                else:
                    anios.append(None)

            for fila in filas[1:]:
                if len(fila) < 2:
                    continue
                
                cuenta_raw = fila[0].strip()
                
                for i_col, valor_str in enumerate(fila[2:]):
                    anio = anios[i_col]
                    if anio is None:
                        continue
                    
                    valor = limpiar_valor(valor_str)
                    cuenta_normalizada = mapear_cuenta_normalizada(cuenta_raw, anio)
                    
                    if anio not in datos_flujo_efectivo:
                        datos_flujo_efectivo[anio] = {}
                    
                    if cuenta_normalizada not in datos_flujo_efectivo[anio]:
                        datos_flujo_efectivo[anio][cuenta_normalizada] = valor
                    elif datos_flujo_efectivo[anio][cuenta_normalizada] == 0 and valor != 0:
                        datos_flujo_efectivo[anio][cuenta_normalizada] = valor
    
    progress_bar.progress((i + 1) / len(archivos))

status_text.empty()
progress_bar.empty()

df_balance = pd.DataFrame.from_dict(datos_balance, orient='index').fillna(0.0).T if datos_balance else pd.DataFrame()
df_resultados = pd.DataFrame.from_dict(datos_resultados, orient='index').fillna(0.0).T if datos_resultados else pd.DataFrame()
df_flujo_efectivo = pd.DataFrame.from_dict(datos_flujo_efectivo, orient='index').fillna(0.0).T if datos_flujo_efectivo else pd.DataFrame()

if not df_balance.empty:
    df_balance = df_balance.reindex(sorted(df_balance.columns), axis=1)

if not df_resultados.empty:
    df_resultados = df_resultados.reindex(sorted(df_resultados.columns), axis=1)

if not df_flujo_efectivo.empty:
    df_flujo_efectivo = df_flujo_efectivo.reindex(sorted(df_flujo_efectivo.columns), axis=1)

# ================= ANÁLISIS VERTICAL Y HORIZONTAL =================
df_vertical_balance = pd.DataFrame()
df_horizontal_balance = pd.DataFrame()

if not df_balance.empty:
    df_vertical_balance = df_balance.copy()
    total_activos_row = None
    for idx in df_vertical_balance.index:
        if "TOTAL" in idx and ("ACTIVO" in idx) and "CORRIENTE" not in idx and "NO CORRIENTE" not in idx:
            total_activos_row = idx
            break
    
    if total_activos_row:
        total_activos = df_vertical_balance.loc[total_activos_row]
        for col in df_vertical_balance.columns:
            if total_activos[col] != 0:
                df_vertical_balance[col] = (df_vertical_balance[col] / total_activos[col]) * 100
        df_vertical_balance = df_vertical_balance.round(2)
    
    df_horizontal_balance = df_balance.copy()
    columnas = df_horizontal_balance.columns.tolist()
    nuevas_columnas = []
    for i in range(len(columnas) - 1):
        anio_actual = columnas[i + 1]
        anio_anterior = columnas[i]
        col_nombre = f"{anio_anterior}-{anio_actual}"
        with pd.option_context('mode.use_inf_as_na', True):
            df_horizontal_balance[col_nombre] = ((df_horizontal_balance[anio_actual] - df_horizontal_balance[anio_anterior]) / 
                                                 df_horizontal_balance[anio_anterior] * 100)
        nuevas_columnas.append(col_nombre)
    df_horizontal_balance = df_horizontal_balance[nuevas_columnas].round(2)
    df_horizontal_balance = df_horizontal_balance.replace([float('inf'), float('-inf')], pd.NA)

df_vertical_resultados = pd.DataFrame()
df_horizontal_resultados = pd.DataFrame()

if not df_resultados.empty:
    df_vertical_resultados = df_resultados.copy()
    ventas_row = buscar_cuenta_flexible(df_resultados, [
        ["INGRESOS", "ACTIVIDADES", "ORDINARIAS"],
        ["VENTAS", "NETAS"]
    ])
    
    if ventas_row:
        ventas = df_vertical_resultados.loc[ventas_row]
        for col in df_vertical_resultados.columns:
            if ventas[col] != 0:
                df_vertical_resultados[col] = (df_vertical_resultados[col] / ventas[col]) * 100
        df_vertical_resultados = df_vertical_resultados.round(2)
    
    df_horizontal_resultados = df_resultados.copy()
    columnas = df_horizontal_resultados.columns.tolist()
    nuevas_columnas = []
    for i in range(len(columnas) - 1):
        anio_actual = columnas[i + 1]
        anio_anterior = columnas[i]
        col_nombre = f"{anio_anterior}-{anio_actual}"
        with pd.option_context('mode.use_inf_as_na', True):
            df_horizontal_resultados[col_nombre] = ((df_horizontal_resultados[anio_actual] - df_horizontal_resultados[anio_anterior]) / 
                                                    df_horizontal_resultados[anio_anterior] * 100)
        nuevas_columnas.append(col_nombre)
    df_horizontal_resultados = df_horizontal_resultados[nuevas_columnas].round(2)
    df_horizontal_resultados = df_horizontal_resultados.replace([float('inf'), float('-inf')], pd.NA)

# ================= CÁLCULO DE RATIOS =================
ratios_data = {}
debug_info = {}
anios_comunes = sorted(list(set(df_balance.columns) & set(df_resultados.columns))) if (not df_balance.empty and not df_resultados.empty) else []

if anios_comunes:
    for i, anio in enumerate(anios_comunes):
        ratios_data[anio] = {}
        debug_info[anio] = {}
        
        # Activo Corriente
        act_corr = buscar_cuenta_flexible(df_balance, [
            ["TOTAL", "ACTIVO", "CORRIENTE"],
            ["TOTAL", "ACTIVOS", "CORRIENTES"]
        ])
        activo_corriente = df_balance.loc[act_corr, anio] if act_corr in df_balance.index else 0.0
        debug_info[anio]["activo_corriente"] = f"{act_corr} = {activo_corriente}"
        
        # Inventarios
        inv = buscar_cuenta_flexible(df_balance, [
            ["INVENTARIOS"],
            ["EXISTENCIAS"]
        ])
        if not inv:
            inv = buscar_cuenta_parcial(df_balance, ["INVENTARIO", "EXISTENCIA"]) if not df_balance.empty else None
        inventarios = df_balance.loc[inv, anio] if inv in df_balance.index else 0.0
        debug_info[anio]["inventarios"] = f"{inv} = {inventarios}"
        
        # Pasivo Corriente
        pas_corr = buscar_cuenta_flexible(df_balance, [
            ["TOTAL", "PASIVO", "CORRIENTE"],
            ["TOTAL", "PASIVOS", "CORRIENTES"]
        ])
        pasivo_corriente = df_balance.loc[pas_corr, anio] if pas_corr in df_balance.index else 0.0
        debug_info[anio]["pasivo_corriente"] = f"{pas_corr} = {pasivo_corriente}"
        
        # Cuentas por Cobrar
        cxc_comerciales = buscar_cuenta_flexible(df_balance, [
            ["CUENTAS", "COBRAR", "COMERCIALES"]
        ])
        if not cxc_comerciales:
            cxc_comerciales = buscar_cuenta_parcial(df_balance, ["CUENTAS", "COBRAR", "COMERCIAL"]) if not df_balance.empty else None
        
        cxc_vinculadas = buscar_cuenta_flexible(df_balance, [
            ["CUENTAS", "COBRAR", "ENTIDADES", "RELACIONADAS"],
            ["CUENTAS", "COBRAR", "VINCULADAS"]
        ])
        if not cxc_vinculadas:
            cxc_vinculadas = buscar_cuenta_parcial(df_balance, ["CUENTAS", "COBRAR", "VINCULADA"]) if not df_balance.empty else None
        
        otras_cxc = buscar_cuenta_flexible(df_balance, [
            ["OTRAS", "CUENTAS", "COBRAR"]
        ])
        
        cxc_val = 0.0
        for cxc_idx in [cxc_comerciales, cxc_vinculadas, otras_cxc]:
            if cxc_idx and cxc_idx in df_balance.index:
                cxc_val += df_balance.loc[cxc_idx, anio]
        debug_info[anio]["cxc"] = f"com:{cxc_comerciales}, vinc:{cxc_vinculadas}, otras:{otras_cxc} = {cxc_val}"
        
        # Activos Totales
        act_tot = buscar_cuenta_flexible(df_balance, [
            ["TOTAL", "ACTIVO"],
            ["TOTAL", "ACTIVOS"]
        ])
        activos_totales = df_balance.loc[act_tot, anio] if act_tot in df_balance.index else 0.0
        debug_info[anio]["activos_totales"] = f"{act_tot} = {activos_totales}"
        
        # Pasivo Total
        pas_tot = buscar_cuenta_flexible(df_balance, [
            ["TOTAL", "PASIVO"],
            ["TOTAL", "PASIVOS"]
        ])
        pasivo_total = df_balance.loc[pas_tot, anio] if pas_tot in df_balance.index else 0.0
        
        # Patrimonio
        patr = buscar_cuenta_flexible(df_balance, [
            ["TOTAL", "PATRIMONIO"],
            ["PATRIMONIO", "NETO"],
            ["TOTAL", "PATRIMONIO", "NETO"]
        ])
        if not patr:
            patr = buscar_cuenta_parcial(df_balance, ["PATRIMONIO"]) if not df_balance.empty else None
        patrimonio = df_balance.loc[patr, anio] if patr in df_balance.index else 0.0
        
        if patrimonio == 0.0 and activos_totales != 0.0:
            patrimonio = activos_totales - pasivo_total
        debug_info[anio]["patrimonio"] = f"{patr} = {patrimonio}"
        
        # Ventas
        ventas = buscar_cuenta_flexible(df_resultados, [
            ["INGRESOS", "ACTIVIDADES", "ORDINARIAS"]
        ])
        if not ventas:
            ventas = buscar_cuenta_parcial(df_resultados, ["INGRESOS", "ACTIVIDADES"]) if not df_resultados.empty else None
        if not ventas:
            ventas = buscar_cuenta_parcial(df_resultados, ["VENTAS", "NETAS"]) if not df_resultados.empty else None
        if not ventas:
            ventas = buscar_cuenta_parcial(df_resultados, ["INGRESOS", "OPERACIONALES"]) if not df_resultados.empty else None
        ventas_val = df_resultados.loc[ventas, anio] if ventas in df_resultados.index else 0.0
        debug_info[anio]["ventas"] = f"{ventas} = {ventas_val}"
        
        # Costo de Ventas
        costo = buscar_cuenta_flexible(df_resultados, [
            ["COSTO", "VENTAS"]
        ])
        if not costo:
            costo = buscar_cuenta_parcial(df_resultados, ["COSTO", "VENTA"]) if not df_resultados.empty else None
        costo_ventas = df_resultados.loc[costo, anio] if costo in df_resultados.index else 0.0
        debug_info[anio]["costo_ventas"] = f"{costo} = {costo_ventas}"
        
        # Utilidad Neta
        util = buscar_cuenta_flexible(df_resultados, [
            ["GANANCIA", "PERDIDA", "NETA", "EJERCICIO"]
        ])
        if not util:
            util = buscar_cuenta_parcial(df_resultados, ["GANANCIA", "NETA", "EJERCICIO"]) if not df_resultados.empty else None
        if not util:
            util = buscar_cuenta_parcial(df_resultados, ["UTILIDAD", "NETA", "EJERCICIO"]) if not df_resultados.empty else None
        if not util and not df_resultados.empty:
            for idx in df_resultados.index:
                if "UTILIDAD" in idx and "EJERCICIO" in idx and "NETA" in idx:
                    util = idx
                    break
        utilidad_neta = df_resultados.loc[util, anio] if util in df_resultados.index else 0.0
        debug_info[anio]["utilidad_neta"] = f"{util} = {utilidad_neta}"
        
        # Promedios
        cxc_prom = cxc_val
        inv_prom = inventarios
        act_prom = activos_totales
        patr_prom = patrimonio

        if i > 0:
            anio_ant = anios_comunes[i-1]

            # CxC anterior
            cxc_ant = 0.0
            for cxc_idx in [cxc_comerciales, cxc_vinculadas, otras_cxc]:
                if cxc_idx and cxc_idx in df_balance.index:
                    cxc_ant += df_balance.loc[cxc_idx, anio_ant]

            # Inventarios anteriores
            inv_ant = df_balance.loc[inv, anio_ant] if inv in df_balance.index else 0.0
            act_ant = df_balance.loc[act_tot, anio_ant] if act_tot in df_balance.index else 0.0

            # Patrimonio anterior
            patr_ant = df_balance.loc[patr, anio_ant] if patr in df_balance.index else 0.0
            if patr_ant == 0.0 and act_ant != 0.0:
                pas_ant = df_balance.loc[pas_tot, anio_ant] if pas_tot in df_balance.index else 0.0
                patr_ant = act_ant - pas_ant

            # Promedios con control
            cxc_prom = (cxc_val + cxc_ant) / 2 if (cxc_val + cxc_ant) != 0 else "N/A"
            inv_prom = (inventarios + inv_ant) / 2 if (inventarios + inv_ant) != 0 else "N/A"
            act_prom = (activos_totales + act_ant) / 2 if (activos_totales + act_ant) != 0 else "N/A"
            patr_prom = (patrimonio + patr_ant) / 2 if (patrimonio + patr_ant) != 0 else "N/A"

        # Calcular ratios con "N/A" si no se puede
        def safe_div(num, den):
            if isinstance(den, (int, float)) and den != 0:
                return num / den
            return "N/A"

        ratios_data[anio]["Liquidez Corriente"] = safe_div(activo_corriente, pasivo_corriente)
        ratios_data[anio]["Prueba Ácida"] = safe_div((activo_corriente - inventarios), pasivo_corriente)
        ratios_data[anio]["Rotación CxC"] = safe_div(ventas_val, cxc_prom) if not isinstance(cxc_prom, str) else "N/A"
        ratios_data[anio]["Rotación Inventarios"] = safe_div(abs(costo_ventas), inv_prom) if not isinstance(inv_prom, str) else "N/A"
        ratios_data[anio]["Rotación Activos Totales"] = safe_div(ventas_val, act_prom) if not isinstance(act_prom, str) else "N/A"
        ratios_data[anio]["Razón Deuda Total"] = safe_div(pasivo_total, activos_totales)
        ratios_data[anio]["Razón Deuda/Patrimonio"] = safe_div(pasivo_total, patrimonio)
        ratios_data[anio]["Margen Neto"] = safe_div(utilidad_neta, ventas_val)
        ratios_data[anio]["ROA"] = safe_div(utilidad_neta, act_prom) if not isinstance(act_prom, str) else "N/A"
        ratios_data[anio]["ROE"] = safe_div(utilidad_neta, patr_prom) if not isinstance(patr_prom, str) else "N/A"

# Crear DataFrame de ratios y redondear sólo celdas numéricas
if ratios_data:
    df_ratios = pd.DataFrame.from_dict(ratios_data, orient='index')
    def round_if_num(x):
        return round(x, 4) if isinstance(x, (int, float)) else x
    df_ratios = df_ratios.applymap(round_if_num).T
else:
    df_ratios = pd.DataFrame()

# ================= SIDEBAR STATUS =================
with st.sidebar:
    st.markdown("---")
    st.success(f"✅ **{len(archivos)}** archivos procesados")
    if anios_comunes:
        st.info(f"📅 **Años:** {', '.join(map(str, anios_comunes))}")
    st.metric("Ratios Calculados", len(df_ratios) if not df_ratios.empty else 0)
    
    # Debug checkbox
    if st.checkbox("🐛 Mostrar Debug Info", value=False):
        st.markdown("### Debug: Cuentas encontradas")
        for anio_debug in anios_comunes:
            if anio_debug in debug_info:
                with st.expander(f"Año {anio_debug}"):
                    for key, val in debug_info[anio_debug].items():
                        st.text(f"{key}: {val}")

# ================= TABS =================
tab1, tab2, tab3, tab4 = st.tabs(["📊 Estados Financieros", "📈 Análisis V/H", "🧮 Ratios y Gráficas", "📥 Descargar"])

with tab1:
    st.subheader("💼 Estado de Situación Financiera")
    if not df_balance.empty:
        st.dataframe(df_balance, use_container_width=True)
    else:
        st.warning("No se encontró data del Balance")
    
    st.markdown("---")
    st.subheader("💰 Estado de Resultados")
    if not df_resultados.empty:
        st.dataframe(df_resultados, use_container_width=True)
    else:
        st.warning("No se encontró data del Estado de Resultados")
    
    st.markdown("---")
    st.subheader("💵 Estado de Flujo de Efectivo")
    if not df_flujo_efectivo.empty:
        st.dataframe(df_flujo_efectivo, use_container_width=True)
    else:
        st.warning("No se encontró data del Flujo de Efectivo")

with tab2:
    st.subheader("📊 Análisis Vertical y Horizontal - Estado de Situación Financiera")
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Análisis Vertical (%)**")
        if not df_vertical_balance.empty:
            st.dataframe(df_vertical_balance, use_container_width=True)
    
    with col2:
        st.markdown("**Análisis Horizontal (Variación %)**")
        if not df_horizontal_balance.empty:
            st.dataframe(df_horizontal_balance, use_container_width=True)
    
    st.markdown("---")
    st.subheader("📊 Análisis Vertical y Horizontal - Estado de Resultados")
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Análisis Vertical (%)**")
        if not df_vertical_resultados.empty:
            st.dataframe(df_vertical_resultados, use_container_width=True)
    
    with col2:
        st.markdown("**Análisis Horizontal (Variación %)**")
        if not df_horizontal_resultados.empty:
            st.dataframe(df_horizontal_resultados, use_container_width=True)

with tab3:
    st.subheader("🧮 Ratios Financieros")
    
    if not df_ratios.empty:
        ultimo_anio = df_ratios.columns[-1]
        penultimo_anio = df_ratios.columns[-2] if len(df_ratios.columns) > 1 else ultimo_anio
        
        # Helper para mostrar métricas con manejo de "N/A"
        def format_pct(val):
            return f"{val:.2%}" if isinstance(val, (int, float)) else "N/A"
        def format_num(val, dec=2):
            return f"{val:.{dec}f}" if isinstance(val, (int, float)) else "N/A"
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            val_actual = df_ratios.loc['ROE', ultimo_anio] if 'ROE' in df_ratios.index else "N/A"
            val_anterior = df_ratios.loc['ROE', penultimo_anio] if 'ROE' in df_ratios.index else "N/A"
            delta = val_actual - val_anterior if isinstance(val_actual,(int,float)) and isinstance(val_anterior,(int,float)) else None
            st.metric("ROE", format_pct(val_actual), delta=(format_pct(delta) if delta is not None else None))
        
        with col2:
            val_actual = df_ratios.loc['ROA', ultimo_anio] if 'ROA' in df_ratios.index else "N/A"
            val_anterior = df_ratios.loc['ROA', penultimo_anio] if 'ROA' in df_ratios.index else "N/A"
            delta = val_actual - val_anterior if isinstance(val_actual,(int,float)) and isinstance(val_anterior,(int,float)) else None
            st.metric("ROA", format_pct(val_actual), delta=(format_pct(delta) if delta is not None else None))
        
        with col3:
            val_actual = df_ratios.loc['Liquidez Corriente', ultimo_anio] if 'Liquidez Corriente' in df_ratios.index else "N/A"
            val_anterior = df_ratios.loc['Liquidez Corriente', penultimo_anio] if 'Liquidez Corriente' in df_ratios.index else "N/A"
            delta = val_actual - val_anterior if isinstance(val_actual,(int,float)) and isinstance(val_anterior,(int,float)) else None
            st.metric("Liquidez Corriente", format_num(val_actual,2), delta=(format_num(delta,2) if delta is not None else None))
        
        with col4:
            val_actual = df_ratios.loc['Margen Neto', ultimo_anio] if 'Margen Neto' in df_ratios.index else "N/A"
            val_anterior = df_ratios.loc['Margen Neto', penultimo_anio] if 'Margen Neto' in df_ratios.index else "N/A"
            delta = val_actual - val_anterior if isinstance(val_actual,(int,float)) and isinstance(val_anterior,(int,float)) else None
            st.metric("Margen Neto", format_pct(val_actual), delta=(format_pct(delta) if delta is not None else None))
        
        st.markdown("---")
        st.markdown("### 📋 Tabla de Ratios")
        st.dataframe(df_ratios, use_container_width=True)
        
        st.markdown("---")
        st.markdown("### 📈 Gráficas Individuales por Ratio")
        
        col1, col2 = st.columns(2)
        for idx, ratio in enumerate(df_ratios.index):
            fig = go.Figure()
            yvals = pd.to_numeric(df_ratios.loc[ratio], errors='coerce')  # convierte "N/A" a NaN
            fig.add_trace(go.Scatter(
                x=df_ratios.columns,
                y=yvals,
                mode='lines+markers',
                name=ratio,
                line=dict(width=3),
                marker=dict(size=8)
            ))
            fig.update_layout(
                title=f"{ratio}",
                xaxis_title="Año",
                yaxis_title="Valor",
                height=350,
                showlegend=False
            )
            if idx % 2 == 0:
                with col1:
                    st.plotly_chart(fig, use_container_width=True)
            else:
                with col2:
                    st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning("No se pudieron calcular ratios")

with tab4:
    st.subheader("📥 Descargar Reporte Consolidado")
    st.markdown(f"**Empresa:** {nombre_empresa}")
    st.markdown(f"**Años analizados:** {', '.join(map(str, anios_comunes)) if anios_comunes else 'N/A'}")
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not df_balance.empty:
            df_balance.to_excel(writer, sheet_name='Balance', index_label='Cuenta')
        if not df_resultados.empty:
            df_resultados.to_excel(writer, sheet_name='Estado Resultados', index_label='Cuenta')
        if not df_flujo_efectivo.empty:
            df_flujo_efectivo.to_excel(writer, sheet_name='Flujo Efectivo', index_label='Cuenta')
        
        if not df_vertical_balance.empty and not df_horizontal_balance.empty:
            df_vertical_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta', startrow=0)
            ws = writer.sheets['Analisis Balance']
            startrow = len(df_vertical_balance) + 3
            ws.cell(row=startrow, column=1, value="ANÁLISIS HORIZONTAL (Variación %)")
            df_horizontal_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta', startrow=startrow+1, header=True)
        elif not df_vertical_balance.empty:
            df_vertical_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta')
        elif not df_horizontal_balance.empty:
            df_horizontal_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta')
        
        if not df_vertical_resultados.empty and not df_horizontal_resultados.empty:
            df_vertical_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta', startrow=0)
            ws = writer.sheets['Analisis Resultados']
            startrow = len(df_vertical_resultados) + 3
            ws.cell(row=startrow, column=1, value="ANÁLISIS HORIZONTAL (Variación %)")
            df_horizontal_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta', startrow=startrow+1, header=True)
        elif not df_vertical_resultados.empty:
            df_vertical_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta')
        elif not df_horizontal_resultados.empty:
            df_horizontal_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta')
        
        if not df_ratios.empty:
            df_ratios.to_excel(writer, sheet_name='Ratios', index_label='Ratio')
    
    output.seek(0)
    wb = load_workbook(output)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(name='Calibri', size=10, bold=True)
    cell_font = Font(name='Calibri', size=10)
    subtitle_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    subtitle_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if isinstance(cell.value, str) and "ANÁLISIS HORIZONTAL" in cell.value:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=cell.row, column=col).fill = subtitle_fill
                        ws.cell(row=cell.row, column=col).font = subtitle_font
                        ws.cell(row=cell.row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=cell.row + 1, column=col).fill = header_fill
                        ws.cell(row=cell.row + 1, column=col).font = header_font
                        ws.cell(row=cell.row + 1, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                
                if isinstance(cell.value, str) and "TOTAL" in cell.value.upper():
                    for c in row:
                        c.fill = total_fill
                        c.font = total_font
                
                if isinstance(cell.value, (int, float)) and cell.column > 1:
                    if 'Analisis' in sheet_name:
                        cell.number_format = '0.0"%"'
                    elif 'Ratios' in sheet_name:
                        cell.number_format = '0.0000'
                    else:
                        cell.number_format = '#,##0'
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # ----------------- APLICAR COLOR SCALE A LAS HOJAS DE ANALISIS -----------------
    color_scale = ColorScaleRule(
        start_type='min', start_color='F8696B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='63BE7B'
    )
    
    # Analisis Balance: aplicar a vertical y horizontal si existen
    if 'Analisis Balance' in wb.sheetnames:
        ws_ab = wb['Analisis Balance']
        if not df_vertical_balance.empty:
            n_rows_v = len(df_vertical_balance)
            n_cols_v = df_vertical_balance.shape[1]
            start_row_v = 2  # encabezado en fila 1
            start_col_v = 2  # datos numéricos desde columna B
            end_row_v = start_row_v + n_rows_v - 1
            end_col_v = start_col_v + n_cols_v - 1
            ws_ab.conditional_formatting.add(f"{get_column_letter(start_col_v)}{start_row_v}:{get_column_letter(end_col_v)}{end_row_v}", color_scale)
        if not df_horizontal_balance.empty:
            startrow_h = len(df_vertical_balance) + 4
            n_rows_h = len(df_horizontal_balance)
            n_cols_h = df_horizontal_balance.shape[1]
            start_col_h = 2
            end_col_h = start_col_h + n_cols_h - 1
            start_row_h = startrow_h + 1
            end_row_h = start_row_h + n_rows_h - 1
            ws_ab.conditional_formatting.add(f"{get_column_letter(start_col_h)}{start_row_h}:{get_column_letter(end_col_h)}{end_row_h}", color_scale)
    
    # Analisis Resultados
    if 'Analisis Resultados' in wb.sheetnames:
        ws_ar = wb['Analisis Resultados']
        if not df_vertical_resultados.empty:
            n_rows_v = len(df_vertical_resultados)
            n_cols_v = df_vertical_resultados.shape[1]
            start_row_v = 2
            start_col_v = 2
            end_row_v = start_row_v + n_rows_v - 1
            end_col_v = start_col_v + n_cols_v - 1
            ws_ar.conditional_formatting.add(f"{get_column_letter(start_col_v)}{start_row_v}:{get_column_letter(end_col_v)}{end_row_v}", color_scale)
        if not df_horizontal_resultados.empty:
            startrow_h = len(df_vertical_resultados) + 4
            n_rows_h = len(df_horizontal_resultados)
            n_cols_h = df_horizontal_resultados.shape[1]
            start_col_h = 2
            end_col_h = start_col_h + n_cols_h - 1
            start_row_h = startrow_h + 1
            end_row_h = start_row_h + n_rows_h - 1
            ws_ar.conditional_formatting.add(f"{get_column_letter(start_col_h)}{start_row_h}:{get_column_letter(end_col_h)}{end_row_h}", color_scale)
    
# ----------------- GRAFICAS EN EXCEL -----------------
if not df_ratios.empty and 'Ratios' in wb.sheetnames:
    ws_ratios = wb['Ratios']
    
    if 'Ratios y Graficas' in wb.sheetnames:
        del wb['Ratios y Graficas']
    ws_graficas = wb.create_sheet('Ratios y Graficas')
    
    ws_graficas['A1'] = 'TABLA DE RATIOS FINANCIEROS'
    ws_graficas['A1'].font = Font(name='Calibri', size=14, bold=True, color="366092")
    ws_graficas.merge_cells('A1:H1')
    ws_graficas.append([])
    
    header_row = ['Ratio / Año'] + [str(y) for y in df_ratios.columns]
    ws_graficas.append(header_row)
    
    for idx, cell in enumerate(ws_graficas[3], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    for ratio_name in df_ratios.index:
        row_data = [ratio_name]
        for col in df_ratios.columns:
            val = df_ratios.loc[ratio_name, col]
            if isinstance(val, (int, float)) and not pd.isna(val):
                row_data.append(val)
            else:
                row_data.append("")
        ws_graficas.append(row_data)
    
    for row_idx in range(4, 4 + len(df_ratios)):
        ws_graficas.cell(row=row_idx, column=1).font = Font(name='Calibri', size=10, bold=True)
        for col_idx in range(2, 2 + len(df_ratios.columns)):
            cell = ws_graficas.cell(row=row_idx, column=col_idx)
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
    ws_graficas.column_dimensions['A'].width = 30
    for col in range(2, len(df_ratios.columns) + 2):
        ws_graficas.column_dimensions[get_column_letter(col)].width = 12

   # ----------------- PREPARAR GRÁFICAS -----------------
    chart_start_row = len(df_ratios) + 6
    ws_graficas.cell(row=chart_start_row, column=1, value='GRÁFICAS INDIVIDUALES POR RATIO')
    ws_graficas.cell(row=chart_start_row, column=1).font = Font(name='Calibri', size=14, bold=True, color="366092")

    chart_row = chart_start_row + 2
    charts_per_row = 2
    chart_height = 15
    chart_width = 10
    num_years = len(df_ratios.columns)

    # ⭐️ Asegurar que los años en la fila 3 sean TEXTO (ideal para BarChart)
    for col in range(2, num_years + 2):
        cell = ws_graficas.cell(row=3, column=col)
        if isinstance(cell.value, (int, float)):
            cell.value = str(int(cell.value))
        elif cell.value is None:
            cell.value = f"Año{col-1}"
        cell.number_format = "@"  # Formato de texto

    # ⭐️ CREAR GRÁFICAS DE BARRAS (UNA POR RATIO)
    for idx, ratio_name in enumerate(df_ratios.index):
        chart = BarChart()
        chart.type = "col"  # Barras verticales
        chart.title = ratio_name
        chart.style = 12
        chart.y_axis.title = "Valor"
        chart.x_axis.title = "Año"
        chart.height = 7.5
        chart.width = 13
        chart.legend = None

        data_row = 4 + idx

        # Referencias de datos
        values = Reference(ws_graficas, 
                        min_col=2,
                        min_row=data_row,
                        max_col=num_years + 1,
                        max_row=data_row)

        categories = Reference(ws_graficas,
                            min_col=2,
                            min_row=3,
                            max_col=num_years + 1,
                            max_row=3)

        # Serie
        series = Series(values, title="")
        chart.series.append(series)
        chart.set_categories(categories)

        # ⭐️ Configurar eje Y
        chart.y_axis.tickLblPos = "low"
        chart.y_axis.majorGridlines = ChartLines()
        chart.y_axis.minorGridlines = None
        chart.y_axis.number_format = '0.00'

        # ⭐️ Configurar eje X
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.majorGridlines = None
        chart.x_axis.minorGridlines = None
        chart.x_axis.number_format = "@"  # Formato de texto

        # ⭐️ Formato del eje Y según los datos
        valores_ratio = []
        for col in range(2, num_years + 2):
            val = ws_graficas.cell(row=data_row, column=col).value
            if isinstance(val, (int, float)):
                valores_ratio.append(val)

        if valores_ratio:
            max_val = max(valores_ratio)
            if abs(max_val) < 1:
                chart.y_axis.number_format = '0.0000'
            elif abs(max_val) < 100:
                chart.y_axis.number_format = '0.00'
            else:
                chart.y_axis.number_format = '#,##0'

            # Forzar rango del eje Y
            min_val = min(valores_ratio)
            chart.y_axis.scaling.min = min_val * 0.9 if min_val > 0 else min_val * 1.1
            chart.y_axis.scaling.max = max_val * 1.1

        # Posición del gráfico
        row_pos = chart_row + (idx // charts_per_row) * chart_height
        col_pos = 1 + (idx % charts_per_row) * chart_width
        cell_pos = f"{get_column_letter(col_pos)}{row_pos}"
        ws_graficas.add_chart(chart, cell_pos)

        print(f"Gráfico {idx + 1}: {ratio_name} en celda {cell_pos}")
        
    output_formatted = io.BytesIO()
    wb.save(output_formatted)
    output_formatted.seek(0)
    
    st.download_button(
        label="📥 Descargar Excel Consolidado (Con Gráficas)",
        data=output_formatted.getvalue(),
        file_name=f"Analisis_Financiero_{nombre_empresa.replace(' ', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.success("✅ ¡Proceso completado! El archivo incluye estados financieros, análisis V/H, ratios y gráficas.")