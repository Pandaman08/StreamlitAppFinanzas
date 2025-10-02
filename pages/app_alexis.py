import streamlit as st
import pandas as pd
import numpy as np
import io
from bs4 import BeautifulSoup
import re
import unicodedata
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule

st.set_page_config(
    page_title="Consolidador Financiero SMV",  
    page_icon="📈",                            
    layout="wide",                            
    initial_sidebar_state="expanded"           
)
# ================= HEADER =================
st.title("📊 Consolidador de Estados Financieros - SMV")
st.markdown("**Análisis Financiero Automatizado** | Sube archivos Excel del SMV y obtén análisis completo con gráficas.")

# ================= SIDEBAR =================
with st.sidebar:
    st.header("⚙️ Configuración del Análisis")
    
    nombre_empresa = st.text_input(
        "Nombre de la Empresa", 
        value="EMPRESA ANALIZADA", 
        help="Este nombre se mostrará en el reporte generado"
    )
    
    st.markdown("---")
    
    st.subheader("📋 Cómo usar la herramienta")
    st.info(
        """
        1️⃣ Descarga los archivos Excel (.xls) desde la SMV.  
        2️⃣ Súbelos aquí (pueden ser de cualquier año).  
        3️⃣ Espera mientras se procesan los datos.  
        4️⃣ Visualiza los resultados y descarga tu consolidado.  
        """
    )
    
    st.markdown("---")
    st.write("💡 Tip: Asegúrate de que los archivos Excel tengan el formato estándar del SMV para un análisis correcto.")

# ================= UPLOAD FILES =================
archivos = st.file_uploader(
    "📁 Selecciona archivos Excel (.xls) del SMV",
    type=["xls"],
    accept_multiple_files=True
)

if not archivos:
    st.warning("👆 **Por favor, sube los archivos Excel del SMV para comenzar el análisis.**")
    st.stop()

# ================= UTILIDADES =================

# Normalizar nombres de cuentas y empresas para comparaciones consistentes.
def normalize_name(s):
    if not isinstance(s, str):
        return s
    s2 = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('ASCII')
    s2 = re.sub(r'\s+', ' ', s2).strip().upper()
    s2 = re.sub(r'\s*\(\d+\)\s*$', '', s2)
    return s2

#Convertir datos de Excel a números válidos para cálculos financieros.
def limpiar_valor(valor):
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    if not valor or valor == '0' or valor == '':
        return 0.0
    valor = str(valor).replace(',', '').replace('\xa0', '').replace(' ', '')
    if valor.startswith('(') and valor.endswith(')'):
        valor = '-' + valor[1:-1]
    try:
        return float(valor)
    except:
        return 0.0

# Buscar cuentas que contengan todas las palabras clave (búsqueda estricta).
def buscar_cuenta_flexible(df, keywords_list):
    for keywords in keywords_list:
        for idx in df.index:
            if all(kw.upper() in idx.upper() for kw in keywords):
                return idx
    return None

# Buscar cuentas que contengan cualquiera de las palabras clave (búsqueda amplia).
def buscar_cuenta_parcial(df, keywords):
    for idx in df.index:
        if any(kw.upper() in idx.upper() for kw in keywords):
            return idx
    return None

# ================= PROCESAR ARCHIVOS =================

#Se crean diccionarios donde se guardaran todos los estados financieros 
datos_balance = {}
datos_resultados = {}
datos_flujo_efectivo = {}

#Barra de progreso y mensaje de estado
progress_bar = st.progress(0)
status_text = st.empty()

#Recorre cada archivo que el usuario subio 
for i, archivo in enumerate(archivos):
    status_text.text(f"📦 Procesando: {archivo.name}")
    
#Leer el contenido del archivo
    contenido = None
    for cod in ['latin-1', 'cp1252', 'utf-8']:
        try:
            archivo.seek(0)
            contenido = archivo.read().decode(cod)
            break
        except:
            continue
#Verificación de lectura
    if not contenido:
        st.error(f"❌ No se pudo leer {archivo.name}")
        continue
    
    soup = BeautifulSoup(contenido, 'html.parser')
    

    # Extraer tabla del balance
    tabla_balance = soup.find('table', {'id': 'gvReporte'})
    if tabla_balance:
        filas = [[td.get_text(strip=True) for td in tr.find_all(['td', 'th'])] for tr in tabla_balance.find_all('tr') if tr.find_all(['td', 'th'])]
        if len(filas) > 1:
            encabezados = filas[0]
            columnas_anios = encabezados[2:]
            anios = [int(m.group(0)) if (m := re.search(r'\b(19|20)\d{2}\b', col)) else None for col in columnas_anios]
            
            for fila in filas[1:]:
                if len(fila) < 3:
                    continue
                cuenta_raw = fila[0].strip()
                if not cuenta_raw:
                    continue
                
                for i_col, valor_str in enumerate(fila[2:]):
                    anio = anios[i_col]
                    if anio is None:
                        continue
                    valor = limpiar_valor(valor_str)
                    cuenta = normalize_name(cuenta_raw)
                    
                    if anio not in datos_balance:
                        datos_balance[anio] = {}
                    
                    if cuenta not in datos_balance[anio]:
                        datos_balance[anio][cuenta] = valor
                    elif datos_balance[anio][cuenta] == 0 and valor != 0:
                        datos_balance[anio][cuenta] = valor
    

    # Estado de Resultados
    tabla_resultados = soup.find('table', {'id': 'gvReporte1'})
    if tabla_resultados:
        filas = [[td.get_text(strip=True) for td in tr.find_all(['td', 'th'])] for tr in tabla_resultados.find_all('tr') if tr.find_all(['td', 'th'])]
        if len(filas) > 1:
            encabezados = filas[0]
            columnas_anios = encabezados[2:]
            anios = [int(m.group(0)) if (m := re.search(r'\b(19|20)\d{2}\b', col)) else None for col in columnas_anios]
            
            for fila in filas[1:]:
                if len(fila) < 3:
                    continue
                cuenta_raw = fila[0].strip()
                
                for i_col, valor_str in enumerate(fila[2:]):
                    anio = anios[i_col]
                    if anio is None:
                        continue
                    valor = limpiar_valor(valor_str)
                    cuenta = normalize_name(cuenta_raw)
                    
                    if anio not in datos_resultados:
                        datos_resultados[anio] = {}
                    datos_resultados[anio][cuenta] = valor
    

    
    # Flujo de Efectivo
    tabla_flujo = soup.find('table', {'id': 'gvReporte3'})
    if tabla_flujo:
        filas = [[td.get_text(strip=True) for td in tr.find_all(['td', 'th'])] for tr in tabla_flujo.find_all('tr') if tr.find_all(['td', 'th'])]
        if len(filas) > 1:
            encabezados = filas[0]
            columnas_anios = encabezados[2:]
            anios = [int(m.group(0)) if (m := re.search(r'\b(19|20)\d{2}\b', col)) else None for col in columnas_anios]
            
            for fila in filas[1:]:
                if len(fila) < 3:
                    continue
                cuenta_raw = fila[0].strip()
                
                for i_col, valor_str in enumerate(fila[2:]):
                    anio = anios[i_col]
                    if anio is None:
                        continue
                    valor = limpiar_valor(valor_str)
                    cuenta = normalize_name(cuenta_raw)
                    
                    if anio not in datos_flujo_efectivo:
                        datos_flujo_efectivo[anio] = {}
                    if cuenta not in datos_flujo_efectivo[anio]:
                        datos_flujo_efectivo[anio][cuenta] = valor
    
    progress_bar.progress((i + 1) / len(archivos))

status_text.empty()
progress_bar.empty()

# ================= CREAR DATAFRAMES =================
df_balance = pd.DataFrame.from_dict(datos_balance, orient='index').fillna(0.0).T if datos_balance else pd.DataFrame()
df_resultados = pd.DataFrame.from_dict(datos_resultados, orient='index').fillna(0.0).T if datos_resultados else pd.DataFrame()
df_flujo_efectivo = pd.DataFrame.from_dict(datos_flujo_efectivo, orient='index').fillna(0.0).T if datos_flujo_efectivo else pd.DataFrame()

if not df_balance.empty:
    df_balance = df_balance.reindex(sorted(df_balance.columns), axis=1)
if not df_resultados.empty:
    df_resultados = df_resultados.reindex(sorted(df_resultados.columns), axis=1)
if not df_flujo_efectivo.empty:
    df_flujo_efectivo = df_flujo_efectivo.reindex(sorted(df_flujo_efectivo.columns), axis=1)

# ================= ANÁLISIS V/H =================
df_vertical_balance = pd.DataFrame()
df_horizontal_balance = pd.DataFrame()

if not df_balance.empty:
    df_vertical_balance = df_balance.copy()
    total_activos_row = None
    for idx in df_vertical_balance.index:
        if "TOTAL" in idx and "ACTIVO" in idx and "CORRIENTE" not in idx and "NO CORRIENTE" not in idx:
            total_activos_row = idx
            break
    
    if total_activos_row:
        total_activos = df_vertical_balance.loc[total_activos_row]
        for col in df_vertical_balance.columns:
            if total_activos[col] != 0:
                df_vertical_balance[col] = (df_vertical_balance[col] / total_activos[col]) * 100
        df_vertical_balance = df_vertical_balance.round(1)
    
    df_horizontal_balance = df_balance.copy()
    columnas = df_horizontal_balance.columns.tolist()
    nuevas_columnas = []
    for i in range(len(columnas) - 1):
        anio_actual = columnas[i + 1]
        anio_anterior = columnas[i]
        col_nombre = f"{anio_anterior}-{anio_actual}"
        df_horizontal_balance[col_nombre] = (df_horizontal_balance[anio_actual] - df_horizontal_balance[anio_anterior]) / df_horizontal_balance[anio_anterior] * 100
        nuevas_columnas.append(col_nombre)
    df_horizontal_balance = df_horizontal_balance[nuevas_columnas].round(1)
    df_horizontal_balance = df_horizontal_balance.replace([np.inf, -np.inf], np.nan)

df_vertical_resultados = pd.DataFrame()
df_horizontal_resultados = pd.DataFrame()

if not df_resultados.empty:
    df_vertical_resultados = df_resultados.copy()
    ventas_row = buscar_cuenta_flexible(df_resultados, [["INGRESOS", "ACTIVIDADES", "ORDINARIAS"], ["VENTAS", "NETAS"]])
    
    if ventas_row:
        ventas = df_vertical_resultados.loc[ventas_row]
        for col in df_vertical_resultados.columns:
            if ventas[col] != 0:
                df_vertical_resultados[col] = (df_vertical_resultados[col] / ventas[col]) * 100
        df_vertical_resultados = df_vertical_resultados.round(1)
    
    df_horizontal_resultados = df_resultados.copy()
    columnas = df_horizontal_resultados.columns.tolist()
    nuevas_columnas = []
    for i in range(len(columnas) - 1):
        anio_actual = columnas[i + 1]
        anio_anterior = columnas[i]
        col_nombre = f"{anio_anterior}-{anio_actual}"
        df_horizontal_resultados[col_nombre] = (df_horizontal_resultados[anio_actual] - df_horizontal_resultados[anio_anterior]) / df_horizontal_resultados[anio_anterior] * 100
        nuevas_columnas.append(col_nombre)
    df_horizontal_resultados = df_horizontal_resultados[nuevas_columnas].round(1)
    df_horizontal_resultados = df_horizontal_resultados.replace([np.inf, -np.inf], np.nan)

# ================= CÁLCULO DE RATIOS CORREGIDO =================
ratios_data = {}
anios_comunes = sorted(list(set(df_balance.columns) & set(df_resultados.columns))) if (not df_balance.empty and not df_resultados.empty) else []

if anios_comunes:
    for i, anio in enumerate(anios_comunes):
        ratios_data[anio] = {}
        
        # BALANCE
        act_corr = buscar_cuenta_flexible(df_balance, [["TOTAL", "ACTIVO", "CORRIENTE"], ["TOTAL", "ACTIVOS", "CORRIENTES"]])
        activo_corriente = df_balance.loc[act_corr, anio] if act_corr else 0.0
        
        inv = buscar_cuenta_flexible(df_balance, [["INVENTARIOS"], ["EXISTENCIAS"]])
        if not inv:
            inv = buscar_cuenta_parcial(df_balance, ["INVENTARIO", "EXISTENCIA"])
        inventarios = df_balance.loc[inv, anio] if inv else 0.0
        
        pas_corr = buscar_cuenta_flexible(df_balance, [["TOTAL", "PASIVO", "CORRIENTE"], ["TOTAL", "PASIVOS", "CORRIENTES"]])
        pasivo_corriente = df_balance.loc[pas_corr, anio] if pas_corr else 0.0
        
        cxc_comerciales = buscar_cuenta_flexible(df_balance, [["CUENTAS", "COBRAR", "COMERCIALES"]])
        if not cxc_comerciales:
            cxc_comerciales = buscar_cuenta_parcial(df_balance, ["CUENTAS", "COBRAR", "COMERCIAL"])
        
        cxc_vinculadas = buscar_cuenta_flexible(df_balance, [["CUENTAS", "COBRAR", "ENTIDADES", "RELACIONADAS"], ["CUENTAS", "COBRAR", "VINCULADAS"]])
        if not cxc_vinculadas:
            cxc_vinculadas = buscar_cuenta_parcial(df_balance, ["CUENTAS", "COBRAR", "VINCULADA"])
        
        otras_cxc = buscar_cuenta_flexible(df_balance, [["OTRAS", "CUENTAS", "COBRAR"]])
        
        cxc_val = 0.0
        for cxc_idx in [cxc_comerciales, cxc_vinculadas, otras_cxc]:
            if cxc_idx and cxc_idx in df_balance.index:
                cxc_val += df_balance.loc[cxc_idx, anio]
        
        act_tot = buscar_cuenta_flexible(df_balance, [["TOTAL", "ACTIVO"], ["TOTAL", "ACTIVOS"]])
        activos_totales = df_balance.loc[act_tot, anio] if act_tot else 0.0
        
        pas_tot = buscar_cuenta_flexible(df_balance, [["TOTAL", "PASIVO"], ["TOTAL", "PASIVOS"]])
        pasivo_total = df_balance.loc[pas_tot, anio] if pas_tot else 0.0
        
        patr = buscar_cuenta_flexible(df_balance, [["TOTAL", "PATRIMONIO"], ["PATRIMONIO", "NETO"]])
        if not patr:
            patr = buscar_cuenta_parcial(df_balance, ["PATRIMONIO"])
        patrimonio = df_balance.loc[patr, anio] if patr else 0.0
        if patrimonio == 0.0 and activos_totales != 0.0:
            patrimonio = activos_totales - pasivo_total
        
        # ESTADO DE RESULTADOS
        ventas = buscar_cuenta_flexible(df_resultados, [["INGRESOS", "ACTIVIDADES", "ORDINARIAS"], ["VENTAS", "NETAS"]])
        if not ventas:
            ventas = buscar_cuenta_parcial(df_resultados, ["VENTAS", "NETAS"])
        if not ventas:
            ventas = buscar_cuenta_parcial(df_resultados, ["INGRESOS", "ACTIVIDADES"])
        ventas_val = df_resultados.loc[ventas, anio] if ventas else 0.0
        
        costo = buscar_cuenta_flexible(df_resultados, [["COSTO", "VENTAS"]])
        if not costo:
            costo = buscar_cuenta_parcial(df_resultados, ["COSTO", "VENTA"])
        costo_ventas = df_resultados.loc[costo, anio] if costo else 0.0
        
        util = buscar_cuenta_flexible(df_resultados, [["GANANCIA", "PERDIDA", "NETA", "EJERCICIO"], ["UTILIDAD", "PERDIDA", "NETA", "EJERCICIO"]])
        if not util:
            util = buscar_cuenta_parcial(df_resultados, ["UTILIDAD", "NETA", "EJERCICIO"])
        if not util:
            util = buscar_cuenta_parcial(df_resultados, ["GANANCIA", "NETA"])
        utilidad_neta = df_resultados.loc[util, anio] if util else 0.0
        
        # PROMEDIOS
        cxc_prom = cxc_val
        inv_prom = inventarios
        act_prom = activos_totales
        patr_prom = patrimonio
        
        if i > 0:
            anio_ant = anios_comunes[i-1]
            cxc_ant = 0.0
            for cxc_idx in [cxc_comerciales, cxc_vinculadas, otras_cxc]:
                if cxc_idx and cxc_idx in df_balance.index:
                    cxc_ant += df_balance.loc[cxc_idx, anio_ant]
            
            inv_ant = df_balance.loc[inv, anio_ant] if inv else 0.0
            act_ant = df_balance.loc[act_tot, anio_ant] if act_tot else 0.0
            patr_ant = df_balance.loc[patr, anio_ant] if patr else 0.0
            if patr_ant == 0.0 and act_ant != 0.0:
                pas_ant = df_balance.loc[pas_tot, anio_ant] if pas_tot else 0.0
                patr_ant = act_ant - pas_ant
            
            cxc_prom = (cxc_val + cxc_ant) / 2 if (cxc_val + cxc_ant) != 0 else cxc_val
            inv_prom = (inventarios + inv_ant) / 2 if (inventarios + inv_ant) != 0 else inventarios
            act_prom = (activos_totales + act_ant) / 2 if (activos_totales + act_ant) != 0 else activos_totales
            patr_prom = (patrimonio + patr_ant) / 2 if (patrimonio + patr_ant) != 0 else patrimonio
        
        # RATIOS CORREGIDOS
        ratios_data[anio]["Liquidez Corriente"] = round(activo_corriente / pasivo_corriente, 4) if pasivo_corriente != 0 else None
        ratios_data[anio]["Prueba Ácida"] = round((activo_corriente - inventarios) / pasivo_corriente, 4) if pasivo_corriente != 0 else None
        ratios_data[anio]["Rotación CxC"] = round(ventas_val / cxc_prom, 4) if cxc_prom != 0 else None
        # CORREGIDO: Sin abs() para mantener negativo
        ratios_data[anio]["Rotación Inventarios"] = round(costo_ventas / inv_prom, 4) if inv_prom != 0 else None
        ratios_data[anio]["Rotación Activos Totales"] = round(ventas_val / act_prom, 4) if act_prom != 0 else None
        ratios_data[anio]["Razón Deuda Total"] = round(pasivo_total / activos_totales, 4) if activos_totales != 0 else None
        ratios_data[anio]["Razón Deuda/Patrimonio"] = round(pasivo_total / patrimonio, 4) if patrimonio != 0 else None
        ratios_data[anio]["Margen Neto"] = round(utilidad_neta / ventas_val, 4) if ventas_val != 0 else None
        # CORREGIDO: ROA y ROE en porcentaje (multiplicar por 100)
        ratios_data[anio]["ROA"] = round((utilidad_neta / act_prom) * 100, 4) if act_prom != 0 else None
        ratios_data[anio]["ROE"] = round((utilidad_neta / patr_prom) * 100, 4) if patr_prom != 0 else None

df_ratios = pd.DataFrame.from_dict(ratios_data, orient='index').T if ratios_data else pd.DataFrame()
if not df_ratios.empty:
    # Crear versión formateada para mostrar
    df_ratios_display = df_ratios.copy()
    for col in df_ratios_display.columns:
        df_ratios_display.loc['ROA', col] = f"{df_ratios.loc['ROA', col]:.2f}%" if pd.notna(df_ratios.loc['ROA', col]) else ""
        df_ratios_display.loc['ROE', col] = f"{df_ratios.loc['ROE', col]:.2f}%" if pd.notna(df_ratios.loc['ROE', col]) else ""
else:
    df_ratios_display = pd.DataFrame()
    
# ================= INTERPRETACIONES DE RATIOS =================
INTERPRETACIONES = {
    "Liquidez Corriente": "Mide la capacidad de pagar deudas a corto plazo. Valores > 1 indican buena liquidez. Valores > 2 pueden indicar recursos ociosos.",
    "Prueba Ácida": "Mide liquidez sin inventarios. Valores > 1 indican capacidad de pago inmediata sin vender inventario.",
    "Rotación CxC": "Veces que se cobran las cuentas al año. Valores altos indican cobros eficientes. Días de cobro = 365/Rotación.",
    "Rotación Inventarios": "Veces que se vende el inventario al año (negativo por costo negativo). Valores altos indican gestión eficiente.",
    "Rotación Activos Totales": "Eficiencia en uso de activos para generar ventas. Valores altos indican mejor aprovechamiento de recursos.",
    "Razón Deuda Total": "Proporción de activos financiados con deuda. Valores < 0.5 indican baja deuda. Valores > 0.7 alto riesgo financiero.",
    "Razón Deuda/Patrimonio": "Relación entre deuda y capital propio. Valores < 1 indican más patrimonio que deuda. Valores > 2 alto apalancamiento.",
    "Margen Neto": "% de utilidad sobre ventas. Valores altos indican buena rentabilidad. Industria intensiva en capital tiene márgenes menores.",
    "ROA": "Rentabilidad sobre activos totales (en %). Valores > 5% considerados buenos. Mide eficiencia en uso de recursos.",
    "ROE": "Rentabilidad sobre patrimonio (en %). Valores > 15% considerados excelentes. Mide retorno para accionistas."
}

# ================= SIDEBAR STATUS =================
with st.sidebar:
    st.markdown("---")
    st.success(f"✅ **{len(archivos)}** archivos procesados")
    if anios_comunes:
        st.info(f"📅 **Años:** {', '.join(map(str, anios_comunes))}")
    st.metric("Ratios Calculados", len(df_ratios) if not df_ratios.empty else 0)

# ================= TABS =================
tab1, tab2, tab3, tab4 = st.tabs(["📊 Estados Financieros", "📈 Análisis V/H", "🧮 Ratios y Gráficas", "📥 Descargar"])

with tab1:
    st.subheader("💼 Estado de Situación Financiera")
    if not df_balance.empty:
        st.dataframe(df_balance, use_container_width=True)
    else:
        st.warning("No se encontró data del Balance")
    
    st.markdown("---")
    st.subheader("💰 Estado de Resultados")
    if not df_resultados.empty:
        st.dataframe(df_resultados, use_container_width=True)
    else:
        st.warning("No se encontró data del Estado de Resultados")
    
    st.markdown("---")
    st.subheader("💵 Estado de Flujo de Efectivo")
    if not df_flujo_efectivo.empty:
        st.dataframe(df_flujo_efectivo, use_container_width=True)
    else:
        st.warning("No se encontró data del Flujo de Efectivo")

with tab2:
    st.subheader("📊 Análisis Vertical y Horizontal - Estado de Situación Financiera")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Análisis Vertical (%)**")
        if not df_vertical_balance.empty:
            st.dataframe(df_vertical_balance, use_container_width=True)
    with col2:
        st.markdown("**Análisis Horizontal (Variación %)**")
        if not df_horizontal_balance.empty:
            st.dataframe(df_horizontal_balance, use_container_width=True)
    
    st.markdown("---")
    st.subheader("📊 Análisis Vertical y Horizontal - Estado de Resultados")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Análisis Vertical (%)**")
        if not df_vertical_resultados.empty:
            st.dataframe(df_vertical_resultados, use_container_width=True)
    with col2:
        st.markdown("**Análisis Horizontal (Variación %)**")
        if not df_horizontal_resultados.empty:
            st.dataframe(df_horizontal_resultados, use_container_width=True)

with tab3:
    st.subheader("🧮 Ratios Financieros")
    if not df_ratios.empty:
        ultimo_anio = df_ratios.columns[-1]
        penultimo_anio = df_ratios.columns[-2] if len(df_ratios.columns) > 1 else ultimo_anio
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            valor = df_ratios.loc['ROE', ultimo_anio] if 'ROE' in df_ratios.index and not pd.isna(df_ratios.loc['ROE', ultimo_anio]) else 0
            valor_ant = df_ratios.loc['ROE', penultimo_anio] if 'ROE' in df_ratios.index and not pd.isna(df_ratios.loc['ROE', penultimo_anio]) else 0
            st.metric("ROE", f"{valor:.2f}%", delta=f"{(valor - valor_ant):.2f}%")
        with col2:
            valor = df_ratios.loc['ROA', ultimo_anio] if 'ROA' in df_ratios.index and not pd.isna(df_ratios.loc['ROA', ultimo_anio]) else 0
            valor_ant = df_ratios.loc['ROA', penultimo_anio] if 'ROA' in df_ratios.index and not pd.isna(df_ratios.loc['ROA', penultimo_anio]) else 0
            st.metric("ROA", f"{valor:.2f}%", delta=f"{(valor - valor_ant):.2f}%")
        with col3:
            valor = df_ratios.loc['Liquidez Corriente', ultimo_anio] if 'Liquidez Corriente' in df_ratios.index and not pd.isna(df_ratios.loc['Liquidez Corriente', ultimo_anio]) else 0
            valor_ant = df_ratios.loc['Liquidez Corriente', penultimo_anio] if 'Liquidez Corriente' in df_ratios.index and not pd.isna(df_ratios.loc['Liquidez Corriente', penultimo_anio]) else 0
            st.metric("Liquidez Corriente", f"{valor:.2f}", delta=f"{(valor - valor_ant):.2f}")
        with col4:
            valor = df_ratios.loc['Margen Neto', ultimo_anio] if 'Margen Neto' in df_ratios.index and not pd.isna(df_ratios.loc['Margen Neto', ultimo_anio]) else 0
            valor_ant = df_ratios.loc['Margen Neto', penultimo_anio] if 'Margen Neto' in df_ratios.index and not pd.isna(df_ratios.loc['Margen Neto', penultimo_anio]) else 0
            st.metric("Margen Neto", f"{valor:.2%}", delta=f"{(valor - valor_ant):.2%}")
        
        st.markdown("---")
        st.markdown("### 📋 Tabla de Ratios")
        st.dataframe(df_ratios_display if not df_ratios_display.empty else df_ratios, use_container_width=True)
        
        st.markdown("---")
        st.markdown("### 🔥 Mapa de Calor")
        fig_heatmap = go.Figure(data=go.Heatmap(
            z=df_ratios.values,
            x=df_ratios.columns,
            y=df_ratios.index,
            colorscale='RdYlGn',
            text=df_ratios.round(2).values,
            texttemplate='%{text}',
            textfont={"size": 10},
            colorbar=dict(title="Valor")
        ))
        fig_heatmap.update_layout(title=f"Mapa de Calor - {nombre_empresa}", xaxis_title="Año", yaxis_title="Ratio", height=500)
        st.plotly_chart(fig_heatmap, use_container_width=True)
        
        st.markdown("---")
        st.markdown("### 📈 Gráficas Individuales")
        col1, col2 = st.columns(2)
        for idx, ratio in enumerate(df_ratios.index):
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=df_ratios.columns, y=df_ratios.loc[ratio], mode='lines+markers', name=ratio, line=dict(width=3), marker=dict(size=10)))
            fig.update_layout(title=f"{ratio}", xaxis_title="Año", yaxis_title="Valor", height=350, showlegend=False)
            if idx % 2 == 0:
                with col1:
                    st.plotly_chart(fig, use_container_width=True)
            else:
                with col2:
                    st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning("No se pudieron calcular ratios")

with tab4:
    st.subheader("📥 Descargar Reporte Consolidado")
    st.markdown(f"**Empresa:** {nombre_empresa}")
    st.markdown(f"**Años:** {', '.join(map(str, anios_comunes)) if anios_comunes else 'N/A'}")
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not df_balance.empty:
            df_balance.to_excel(writer, sheet_name='Balance', index_label='Cuenta')
        if not df_resultados.empty:
            df_resultados.to_excel(writer, sheet_name='Estado Resultados', index_label='Cuenta')
        if not df_flujo_efectivo.empty:
            df_flujo_efectivo.to_excel(writer, sheet_name='Flujo Efectivo', index_label='Cuenta')
        
        if not df_vertical_balance.empty and not df_horizontal_balance.empty:
            df_vertical_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta', startrow=0)
            ws = writer.sheets['Analisis Balance']
            startrow = len(df_vertical_balance) + 3
            ws.cell(row=startrow, column=1, value="ANÁLISIS HORIZONTAL")
            df_horizontal_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta', startrow=startrow+1, header=True)
        
        if not df_vertical_resultados.empty and not df_horizontal_resultados.empty:
            df_vertical_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta', startrow=0)
            ws = writer.sheets['Analisis Resultados']
            startrow = len(df_vertical_resultados) + 3
            ws.cell(row=startrow, column=1, value="ANÁLISIS HORIZONTAL")
            df_horizontal_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta', startrow=startrow+1, header=True)
        
        if not df_ratios.empty:
            df_ratios.to_excel(writer, sheet_name='Ratios', index_label='Ratio')
    
    output.seek(0)
    wb = load_workbook(output)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(name='Calibri', size=10, bold=True)
    cell_font = Font(name='Calibri', size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                if isinstance(cell.value, str) and "TOTAL" in cell.value.upper():
                    for c in row:
                        c.fill = total_fill
                        c.font = total_font
                if isinstance(cell.value, (int, float)) and cell.column > 1:
                    if 'Analisis' in sheet_name:
                        cell.number_format = '0.0"%"'
                    elif 'Ratios' in sheet_name:
                        cell.number_format = '0.0000'
                    else:
                        cell.number_format = '#,##0'
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # GRÁFICAS CORREGIDAS
    if not df_ratios.empty and 'Ratios' in wb.sheetnames:
        ws_ratios = wb['Ratios']
        
        if 'Ratios y Graficas' in wb.sheetnames:
            del wb['Ratios y Graficas']
        ws_graficas = wb.create_sheet('Ratios y Graficas')
        
        # Mapa de calor
        ws_graficas['A1'] = 'MAPA DE CALOR DE RATIOS'
        ws_graficas['A1'].font = Font(name='Calibri', size=14, bold=True, color="366092")
        ws_graficas.append([])
        ws_graficas.append(['Ratio'] + df_ratios.columns.tolist())
        
        # Reemplazar None con 0 para gráficas
        df_ratios_clean = df_ratios.fillna(0)
        
        for ratio in df_ratios_clean.index:
            row_data = [ratio] + df_ratios_clean.loc[ratio].tolist()
            ws_graficas.append(row_data)
        
        num_years = len(df_ratios_clean.columns)
        ratio_names = df_ratios_clean.index.tolist()
        
        color_scale = ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84', end_type='max', end_color='63BE7B')
        ws_graficas.conditional_formatting.add(f'B3:{get_column_letter(num_years+1)}{len(ratio_names)+2}', color_scale)
        
        for cell in ws_graficas[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in ws_graficas.iter_rows(min_row=3, max_row=len(ratio_names)+2, min_col=1, max_col=num_years+1):
            for cell in row:
                cell.font = cell_font
                cell.alignment = Alignment(horizontal='center')
                if cell.column > 1:
                    cell.number_format = '0.0000'
        
        ws_graficas.column_dimensions['A'].width = 30
        for col in range(2, num_years + 2):
            ws_graficas.column_dimensions[get_column_letter(col)].width = 12
        
        # Gráficas con interpretaciones
        chart_start = len(ratio_names) + 5
        ws_graficas.cell(row=chart_start, column=1, value='GRÁFICAS CON INTERPRETACIONES')
        ws_graficas.cell(row=chart_start, column=1).font = Font(name='Calibri', size=14, bold=True, color="366092")
        
        current_row = chart_start + 2
        for idx, ratio in enumerate(ratio_names):
            # Interpretación
            ws_graficas.cell(row=current_row, column=1, value=f"{ratio}:")
            ws_graficas.cell(row=current_row, column=1).font = Font(name='Calibri', size=11, bold=True)
            ws_graficas.cell(row=current_row + 1, column=1, value=INTERPRETACIONES.get(ratio, ""))
            ws_graficas.cell(row=current_row + 1, column=1).font = Font(name='Calibri', size=10, italic=True)
            ws_graficas.merge_cells(f'A{current_row+1}:H{current_row+1}')
            
            # Gráfica
            chart = LineChart()
            chart.title = ratio
            chart.style = 10
            chart.y_axis.title = "Valor"
            chart.x_axis.title = "Año"
            chart.height = 7
            chart.width = 14
            
            cats = Reference(ws_ratios, min_col=2, min_row=1, max_col=num_years+1, max_row=1)
            ratio_row_in_ratios = ratio_names.index(ratio) + 2
            data = Reference(ws_ratios, min_col=2, min_row=ratio_row_in_ratios, max_col=num_years+1, max_row=ratio_row_in_ratios)
            
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            
            ws_graficas.add_chart(chart, f'B{current_row + 3}')
            current_row += 18
    
    output_formatted = io.BytesIO()
    wb.save(output_formatted)
    output_formatted.seek(0)
    
    st.download_button(
        label="📥 Descargar Excel Consolidado",
        data=output_formatted.getvalue(),
        file_name=f"Analisis_{nombre_empresa.replace(' ', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.success("✅ Excel con gráficas e interpretaciones generado")