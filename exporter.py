import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series, ScatterChart
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart.axis import ChartLines
import matplotlib.pyplot as plt
import pandas as pd

def exportar_a_excel(df_balance, df_resultados, df_flujo_efectivo, df_vertical_balance, df_horizontal_balance, df_vertical_resultados, df_horizontal_resultados, df_ratios, nombre_empresa, anios_comunes):
    """Exporta todos los datos a un archivo Excel con estilos y gráficas."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not df_balance.empty:
            df_balance.to_excel(writer, sheet_name='Balance', index_label='Cuenta')
        if not df_resultados.empty:
            df_resultados.to_excel(writer, sheet_name='Estado Resultados', index_label='Cuenta')
        if not df_flujo_efectivo.empty:
            df_flujo_efectivo.to_excel(writer, sheet_name='Flujo Efectivo', index_label='Cuenta')
        if not df_vertical_balance.empty and not df_horizontal_balance.empty:
            df_vertical_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta', startrow=0)
            ws = writer.sheets['Analisis Balance']
            startrow = len(df_vertical_balance) + 3
            ws.cell(row=startrow, column=1, value="ANÁLISIS HORIZONTAL (Variación %)")
            df_horizontal_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta', startrow=startrow+1, header=True)
        elif not df_vertical_balance.empty:
            df_vertical_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta')
        elif not df_horizontal_balance.empty:
            df_horizontal_balance.to_excel(writer, sheet_name='Analisis Balance', index_label='Cuenta')
        if not df_vertical_resultados.empty and not df_horizontal_resultados.empty:
            df_vertical_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta', startrow=0)
            ws = writer.sheets['Analisis Resultados']
            startrow = len(df_vertical_resultados) + 3
            ws.cell(row=startrow, column=1, value="ANÁLISIS HORIZONTAL (Variación %)")
            df_horizontal_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta', startrow=startrow+1, header=True)
        elif not df_vertical_resultados.empty:
            df_vertical_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta')
        elif not df_horizontal_resultados.empty:
            df_horizontal_resultados.to_excel(writer, sheet_name='Analisis Resultados', index_label='Cuenta')
        if not df_ratios.empty:
            df_ratios.to_excel(writer, sheet_name='Ratios', index_label='Ratio')

    output.seek(0)
    wb = load_workbook(output)

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(name='Calibri', size=10, bold=True)
    cell_font = Font(name='Calibri', size=10)
    subtitle_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    subtitle_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if isinstance(cell.value, str) and "ANÁLISIS HORIZONTAL" in cell.value:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=cell.row, column=col).fill = subtitle_fill
                        ws.cell(row=cell.row, column=col).font = subtitle_font
                        ws.cell(row=cell.row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=cell.row + 1, column=col).fill = header_fill
                        ws.cell(row=cell.row + 1, column=col).font = header_font
                        ws.cell(row=cell.row + 1, column=col).alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                if isinstance(cell.value, str) and "TOTAL" in cell.value.upper():
                    for c in row:
                        c.fill = total_fill
                        c.font = total_font
                if isinstance(cell.value, (int, float)) and cell.column > 1:
                    if 'Analisis' in sheet_name:
                        cell.number_format = '0.0"%"'
                    elif 'Ratios' in sheet_name:
                        cell.number_format = '0.0000'
                    else:
                        cell.number_format = '#,##0'
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Color scale
    color_scale = ColorScaleRule(
        start_type='min', start_color='F8696B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='63BE7B'
    )
    if 'Analisis Balance' in wb.sheetnames:
        ws_ab = wb['Analisis Balance']
        if not df_vertical_balance.empty:
            n_rows_v = len(df_vertical_balance)
            n_cols_v = df_vertical_balance.shape[1]
            start_row_v = 2
            start_col_v = 2
            end_row_v = start_row_v + n_rows_v - 1
            end_col_v = start_col_v + n_cols_v - 1
            ws_ab.conditional_formatting.add(f"{get_column_letter(start_col_v)}{start_row_v}:{get_column_letter(end_col_v)}{end_row_v}", color_scale)
        if not df_horizontal_balance.empty:
            startrow_h = len(df_vertical_balance) + 4
            n_rows_h = len(df_horizontal_balance)
            n_cols_h = df_horizontal_balance.shape[1]
            start_col_h = 2
            end_col_h = start_col_h + n_cols_h - 1
            start_row_h = startrow_h + 1
            end_row_h = start_row_h + n_rows_h - 1
            ws_ab.conditional_formatting.add(f"{get_column_letter(start_col_h)}{start_row_h}:{get_column_letter(end_col_h)}{end_row_h}", color_scale)
    if 'Analisis Resultados' in wb.sheetnames:
        ws_ar = wb['Analisis Resultados']
        if not df_vertical_resultados.empty:
            n_rows_v = len(df_vertical_resultados)
            n_cols_v = df_vertical_resultados.shape[1]
            start_row_v = 2
            start_col_v = 2
            end_row_v = start_row_v + n_rows_v - 1
            end_col_v = start_col_v + n_cols_v - 1
            ws_ar.conditional_formatting.add(f"{get_column_letter(start_col_v)}{start_row_v}:{get_column_letter(end_col_v)}{end_row_v}", color_scale)
        if not df_horizontal_resultados.empty:
            startrow_h = len(df_vertical_resultados) + 4
            n_rows_h = len(df_horizontal_resultados)
            n_cols_h = df_horizontal_resultados.shape[1]
            start_col_h = 2
            end_col_h = start_col_h + n_cols_h - 1
            start_row_h = startrow_h + 1
            end_row_h = start_row_h + n_rows_h - 1
            ws_ar.conditional_formatting.add(f"{get_column_letter(start_col_h)}{start_row_h}:{get_column_letter(end_col_h)}{end_row_h}", color_scale)

    # Gráficas en Excel
    if not df_ratios.empty and 'Ratios' in wb.sheetnames:
        ws_ratios = wb['Ratios']
        if 'Ratios y Graficas' in wb.sheetnames:
            del wb['Ratios y Graficas']
        ws_graficas = wb.create_sheet('Ratios y Graficas')
        ws_graficas['A1'] = 'TABLA DE RATIOS FINANCIEROS'
        ws_graficas['A1'].font = Font(name='Calibri', size=14, bold=True, color="366092")
        ws_graficas.merge_cells('A1:H1')
        ws_graficas.append([])
        header_row = ['Ratio / Año'] + [str(y) for y in df_ratios.columns]
        ws_graficas.append(header_row)
        for idx, cell in enumerate(ws_graficas[3], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        for ratio_name in df_ratios.index:
            row_data = [ratio_name]
            for col in df_ratios.columns:
                val = df_ratios.loc[ratio_name, col]
                if isinstance(val, (int, float)) and not pd.isna(val):
                    row_data.append(val)
                else:
                    row_data.append("")
            ws_graficas.append(row_data)
        for row_idx in range(4, 4 + len(df_ratios)):
            ws_graficas.cell(row=row_idx, column=1).font = Font(name='Calibri', size=10, bold=True)
            for col_idx in range(2, 2 + len(df_ratios.columns)):
                cell = ws_graficas.cell(row=row_idx, column=col_idx)
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
        ws_graficas.column_dimensions['A'].width = 30
        for col in range(2, len(df_ratios.columns) + 2):
            ws_graficas.column_dimensions[get_column_letter(col)].width = 12

        chart_start_row = len(df_ratios) + 6
        ws_graficas.cell(row=chart_start_row, column=1, value='GRÁFICAS INDIVIDUALES POR RATIO')
        ws_graficas.cell(row=chart_start_row, column=1).font = Font(name='Calibri', size=14, bold=True, color="366092")
        chart_row = chart_start_row + 2
        charts_per_row = 2
        chart_height = 20
        chart_width = 10
        from openpyxl.drawing.image import Image as XLImage
        from io import BytesIO

        for idx, ratio_name in enumerate(df_ratios.index):
            years = []
            values = []
            for col in df_ratios.columns:
                val = df_ratios.loc[ratio_name, col]
                if isinstance(val, (int, float)) and not pd.isna(val):
                    years.append(str(col))
                    values.append(val)
            if not years or not values:
                continue
            plt.figure(figsize=(6, 4))
            plt.plot(years, values, marker='o', linewidth=2, markersize=6, color='#007acc')
            plt.title(ratio_name, fontsize=12, fontweight='bold')
            plt.xlabel('Año', fontsize=10)
            plt.ylabel('Valor', fontsize=10)
            plt.grid(True, axis='y', linestyle='--', alpha=0.7)
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            img_buffer.seek(0)
            img = XLImage(img_buffer)
            img.width = 400
            img.height = 250
            row_pos = chart_row + (idx // charts_per_row) * chart_height
            col_pos = 1 + (idx % charts_per_row) * chart_width
            cell_pos = f"{get_column_letter(col_pos)}{row_pos}"
            ws_graficas.add_image(img, cell_pos)

    output_formatted = io.BytesIO()
    wb.save(output_formatted)
    output_formatted.seek(0)
    return output_formatted